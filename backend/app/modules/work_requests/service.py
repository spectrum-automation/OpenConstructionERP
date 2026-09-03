# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
"""Work Requests business logic - where the rails are enforced.

Every rule below is refused SERVER-SIDE. The UI mirrors them for a good
experience, but the UI is not the rail: a rail enforced in one code path
is not a rail.

The rails:
    * a reference is minted under a row lock and never re-issued;
    * the status machine refuses an illegal move and says what IS allowed;
    * a stage with ``closes`` is the only thing (besides an explicit
      status change) that marks a request complete;
    * only the requester or a manager closes;
    * the requester, the department (lead, members, assignees, the
      responsible person) or a manager may change a request - nobody
      else, however friendly;
    * every mutation writes its own activity line, server-side.
"""

from __future__ import annotations

import logging
import re
import uuid
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy import Select, String, cast, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.work_requests.models import (
    ACTIVE_STATUSES,
    BALL_IN_COURT,
    COMMENT_KINDS,
    DONE_STATUSES,
    FIELD_TYPES,
    MAX_ASSIGNEES,
    MAX_BOQ_POSITIONS,
    MAX_BULK_IDS,
    MAX_CHECKLIST_ITEMS,
    MAX_COMMENT_CHARS,
    MAX_EXPORT_ROWS,
    MAX_LINKS,
    MAX_REQUEST_TYPES,
    MAX_TEXT_CHARS,
    MAX_TITLE_CHARS,
    PRIORITIES,
    STATUSES,
    TERMINAL_STATUSES,
    WorkDepartment,
    WorkPlannerAlloc,
    WorkPlannerCapacity,
    WorkRequest,
    WorkRequestComment,
    WorkRequestCounter,
    WorkRequestHours,
)

logger = logging.getLogger(__name__)


# ── Errors ───────────────────────────────────────────────────────────────


class WorkRequestError(Exception):
    """Router-facing refusal with a message the person can act on (400)."""


class NotFoundError(WorkRequestError):
    """The thing named does not exist (404)."""


class NotPermitted(WorkRequestError):
    """The caller may not do this to this request (403)."""


class TransitionError(WorkRequestError):
    """An illegal status move. Carries what IS allowed from here (409)."""

    def __init__(self, message: str, allowed: list[str]) -> None:
        super().__init__(message)
        self.allowed = allowed


class ConflictError(WorkRequestError):
    """The record is not in a state where this action makes sense (409)."""


# ── Vocabulary ───────────────────────────────────────────────────────────

#: House prefix on every reference: ``WR-ENG-000001``.
REFERENCE_HOUSE = "WR"
REFERENCE_DIGITS = 6

#: What each status may move to. ``cancelled`` is reachable from every
#: open state; ``closed`` only from ``complete``; nothing leaves a
#: terminal state.
TRANSITIONS: dict[str, tuple[str, ...]] = {
    "draft": ("submitted", "cancelled"),
    "submitted": ("accepted", "cancelled"),
    "accepted": ("in_progress", "on_hold", "cancelled"),
    "in_progress": ("on_hold", "review", "complete", "cancelled"),
    "on_hold": ("accepted", "in_progress", "review", "cancelled"),
    "review": ("in_progress", "complete", "cancelled"),
    "complete": ("closed", "in_progress"),
    "closed": (),
    "cancelled": (),
}

_ISO_DAY = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_KEY_RX = re.compile(r"^[a-z][a-z0-9_]{0,39}$")
_UUID_RX = re.compile(r"^[0-9a-fA-F-]{36}$")
_COLOURS = frozenset(
    {"slate", "blue", "teal", "green", "amber", "orange", "rose", "violet", "red", "grey", "gray", "indigo", "cyan"}
)

#: PostgreSQL text columns reject NUL outright, and a stray control
#: character is never meaningful. Tab, newline and CR are KEPT - a pasted
#: table is made of them.
_CONTROL_STRIP = {c: None for c in range(32) if c not in (9, 10, 13)}
_CONTROL_STRIP[127] = None


def clean_text(value: Any) -> Any:
    """Strip control characters, recursing into containers."""
    if isinstance(value, str):
        return value.translate(_CONTROL_STRIP)
    if isinstance(value, list):
        return [clean_text(v) for v in value]
    if isinstance(value, dict):
        return {clean_text(k): clean_text(v) for k, v in value.items()}
    return value


def single_line(value: Any) -> str:
    return " ".join(str(value or "").split())


def _now() -> datetime:
    return datetime.now(UTC)


def _today() -> date:
    """The calendar day due dates are judged against.

    Server-LOCAL, like team_standup - a due date is a label on the wall
    calendar of the people doing the work, and a UTC day would call a
    board due "today" overdue-or-not by the wrong day for the whole
    morning in any timezone east of Greenwich.
    """
    return date.today()


def _iso(value: Any, *, field: str) -> str | None:
    """A calendar day as ``YYYY-MM-DD`` or None. Anything else is refused."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()[:10]
    if not _ISO_DAY.match(text):
        raise WorkRequestError(f"{field} must be a date (YYYY-MM-DD), not {value!r}")
    try:
        date.fromisoformat(text)
    except ValueError:
        raise WorkRequestError(f"{field} is not a real date: {value!r}") from None
    return text


def _hours(value: Any, *, field: str, allow_none: bool = True) -> float | None:
    if value is None or value == "":
        if allow_none:
            return None
        raise WorkRequestError(f"{field} is required")
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise WorkRequestError(f"{field} must be a number of hours, not {value!r}") from None
    if number != number or number in (float("inf"), float("-inf")):  # NaN / inf
        raise WorkRequestError(f"{field} must be a finite number of hours")
    if number < 0:
        raise WorkRequestError(f"{field} cannot be negative")
    if number > 100_000:
        raise WorkRequestError(f"{field} is implausibly large ({number})")
    return round(number, 2)


def _money_text(value: Any, *, field: str) -> str | None:
    """Money as normalised text (``"185.00"``), never a float."""
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().replace(",", "").lstrip("$")
    try:
        amount = Decimal(text)
    except InvalidOperation:
        raise WorkRequestError(f"{field} must be an amount, not {value!r}") from None
    if amount < 0:
        raise WorkRequestError(f"{field} cannot be negative")
    return f"{amount.quantize(Decimal('0.01'))}"


def _as_uuid(value: Any) -> uuid.UUID | None:
    try:
        return uuid.UUID(str(value))
    except (ValueError, TypeError, AttributeError):
        return None


def _like_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


# ── People ───────────────────────────────────────────────────────────────


async def user_names(session: AsyncSession, ids: set[str] | list[str]) -> dict[str, str]:
    """``{user id: display name}`` for every id that names a real user."""
    from app.modules.users.models import User

    wanted = [u for u in {_as_uuid(i) for i in ids} if u is not None]
    if not wanted:
        return {}
    rows = (await session.execute(select(User.id, User.full_name, User.email).where(User.id.in_(wanted)))).all()
    return {str(uid): (name or email or "") for uid, name, email in rows}


async def _user_name(session: AsyncSession, user_id: str | None) -> str:
    if not user_id:
        return ""
    return (await user_names(session, {str(user_id)})).get(str(user_id), "")


async def _validate_users(session: AsyncSession, ids: list[str] | None, *, field: str) -> list[str]:
    """Unique, real, ACTIVE user ids in the order given."""
    from app.modules.users.models import User

    seen: list[str] = []
    for raw in ids or []:
        text = str(raw or "").strip()
        if text and text not in seen:
            seen.append(text)
    if len(seen) > MAX_ASSIGNEES:
        raise WorkRequestError(f"{field}: at most {MAX_ASSIGNEES} people")
    if not seen:
        return []
    as_uuid = [_as_uuid(s) for s in seen]
    if any(u is None for u in as_uuid):
        raise WorkRequestError(f"{field}: every entry must be a user id")
    rows = (
        (
            await session.execute(select(User.id).where(User.id.in_(as_uuid), User.is_active == True))  # noqa: E712
        )
        .scalars()
        .all()
    )
    found = {str(r) for r in rows}
    missing = [s for s in seen if s not in found]
    if missing:
        raise WorkRequestError(f"{field}: unknown or inactive user {missing[0]}")
    return seen


# ── Departments ──────────────────────────────────────────────────────────


async def list_departments(session: AsyncSession, *, include_inactive: bool = True) -> list[WorkDepartment]:
    stmt = select(WorkDepartment).order_by(WorkDepartment.position, WorkDepartment.name)
    if not include_inactive:
        stmt = stmt.where(WorkDepartment.active == True)  # noqa: E712
    return list((await session.execute(stmt)).scalars().all())


async def ensure_seeded(session: AsyncSession) -> None:
    """First read on a fresh install plants the defaults (idempotent)."""
    from app.modules.work_requests.seeds import seed_departments_if_empty

    await seed_departments_if_empty(session)


async def department_or_error(session: AsyncSession, key: str) -> WorkDepartment:
    row = (
        await session.execute(select(WorkDepartment).where(WorkDepartment.key == str(key or "").strip().lower()))
    ).scalar_one_or_none()
    if row is None:
        raise NotFoundError(f"Unknown department {key!r}")
    return row


async def _departments_by_key(session: AsyncSession) -> dict[str, WorkDepartment]:
    return {d.key: d for d in await list_departments(session)}


def ordered_request_types(d: WorkDepartment, *, include_inactive: bool = False) -> list[dict]:
    """The department's types by ``position`` - live ones only by default.

    A retired type is still returned to a manage caller (and stays
    readable on the requests that already carry it), never offered on a
    raise form.
    """
    out = [dict(t) for t in (d.request_types or [])]
    for i, t in enumerate(out):
        t.setdefault("active", True)
        # A type stored before checklists existed reads back with an empty
        # one, so no caller has to know which release wrote it.
        t.setdefault("checklist", [])
        t["active"] = bool(t["active"])
        t["position"] = int(t.get("position") if t.get("position") is not None else i)
    out.sort(key=lambda t: (t["position"], str(t.get("key") or "")))
    if not include_inactive:
        out = [t for t in out if t["active"]]
    return out


def department_payload(d: WorkDepartment, *, include_inactive: bool = False) -> dict[str, Any]:
    return {
        "key": d.key,
        "name": d.name,
        "prefix": d.prefix,
        "colour": d.colour,
        "description": d.description or "",
        "active": bool(d.active),
        "position": int(d.position or 0),
        "lead_user_id": d.lead_user_id,
        "member_ids": list(d.member_ids or []),
        "hourly_rate": d.hourly_rate,
        "target_days": d.target_days,
        "stages": [dict(s) for s in (d.stages or [])],
        "request_types": ordered_request_types(d, include_inactive=include_inactive),
    }


def _target_days(value: Any) -> int | None:
    """A turnaround target in working days, or None for "no target"."""
    if value is None or str(value).strip() == "":
        return None
    try:
        days = int(value)
    except (TypeError, ValueError):
        raise WorkRequestError(f"target_days must be a whole number of working days, not {value!r}") from None
    if days < 0:
        raise WorkRequestError("target_days cannot be negative")
    if days > 365:
        raise WorkRequestError("target_days is implausibly large (at most 365)")
    return days


def _normalise_stages(stages: list[dict] | None) -> list[dict]:
    out: list[dict] = []
    seen: set[str] = set()
    for i, raw in enumerate(stages or []):
        key = str(raw.get("key") or "").strip().lower()
        name = single_line(clean_text(raw.get("name") or ""))[:120]
        if not _KEY_RX.match(key):
            raise WorkRequestError(f"Stage key {key!r} must be lowercase letters, digits and underscores")
        if key in seen:
            raise WorkRequestError(f"Stage key {key!r} appears twice")
        if not name:
            raise WorkRequestError(f"Stage {key!r} needs a name")
        colour = str(raw.get("colour") or "slate").strip().lower()
        if colour not in _COLOURS:
            raise WorkRequestError(f"Unknown colour {colour!r}")
        seen.add(key)
        order = raw.get("order")
        out.append(
            {
                "key": key,
                "name": name,
                "colour": colour,
                "order": int(order) if order is not None else i,
                "closes": bool(raw.get("closes", False)),
            }
        )
    out.sort(key=lambda s: (s["order"], s["key"]))
    for i, s in enumerate(out):
        s["order"] = i
    return out


def _normalise_field(raw: dict) -> dict:
    key = str(raw.get("key") or "").strip().lower()
    if not _KEY_RX.match(key):
        raise WorkRequestError(f"Field key {key!r} must be lowercase letters, digits and underscores")
    ftype = str(raw.get("type") or "text").strip().lower()
    if ftype not in FIELD_TYPES:
        raise WorkRequestError(f"Field {key!r}: unknown type {ftype!r} (one of {', '.join(FIELD_TYPES)})")
    label = single_line(clean_text(raw.get("label") or key))[:200]
    options = [single_line(clean_text(o))[:200] for o in (raw.get("options") or []) if str(o).strip()]
    if ftype == "select" and not options:
        raise WorkRequestError(f"Select field {key!r} needs options")
    out: dict = {"key": key, "label": label, "type": ftype, "required": bool(raw.get("required", False))}
    if ftype == "select":
        out["options"] = options
    return out


def _normalise_checklist_item(raw: Any) -> dict:
    """One ``{key, label, required}`` - validated exactly like a field.

    A bare string is taken as a label and slugged, so a manage screen can
    offer a plain list box without inventing keys for the operator.
    """
    if isinstance(raw, str):
        raw = {"label": raw}
    if not isinstance(raw, dict):
        raise WorkRequestError(f"A checklist item is a label or {{key, label}}, not {raw!r}")
    label = single_line(clean_text(raw.get("label") or ""))[:200]
    key = str(raw.get("key") or "").strip().lower() or slugify_key(label)
    if not _KEY_RX.match(key):
        raise WorkRequestError(f"Checklist key {key!r} must be lowercase letters, digits and underscores")
    return {"key": key, "label": label or key, "required": bool(raw.get("required", False))}


def _normalise_checklist(items: Any) -> list[dict]:
    """The type's checklist, keys unique - the same rail ``fields`` has."""
    if items is None:
        return []
    if isinstance(items, (str, bytes, dict)):
        raise WorkRequestError("checklist must be a list of items")
    out: list[dict] = []
    seen: set[str] = set()
    for raw in items:
        item = _normalise_checklist_item(raw)
        if item["key"] in seen:
            raise WorkRequestError(f"Checklist key {item['key']!r} appears twice")
        seen.add(item["key"])
        out.append(item)
    if len(out) > MAX_CHECKLIST_ITEMS:
        raise WorkRequestError(f"At most {MAX_CHECKLIST_ITEMS} checklist items on one request type")
    return out


def _normalise_request_type(raw: dict, *, position: int) -> dict:
    key = str(raw.get("key") or "").strip().lower()
    if not _KEY_RX.match(key):
        raise WorkRequestError(f"Request type key {key!r} must be lowercase letters, digits and underscores")
    label = single_line(clean_text(raw.get("label") or key))[:200]
    disciplines = []
    for d in raw.get("disciplines") or []:
        text = str(d or "").strip().lower()
        if text and text not in disciplines:
            disciplines.append(text[:40])
    fields = [_normalise_field(f) for f in (raw.get("fields") or [])]
    fkeys = [f["key"] for f in fields]
    if len(fkeys) != len(set(fkeys)):
        raise WorkRequestError(f"Request type {key!r} has a duplicate field key")
    given = raw.get("position")
    return {
        "key": key,
        "label": label,
        "disciplines": disciplines,
        "fields": fields,
        "checklist": _normalise_checklist(raw.get("checklist")),
        "active": bool(raw.get("active", True)),
        "position": int(given) if given is not None else position,
    }


def _renumber(types: list[dict]) -> list[dict]:
    """Sorted by position, then numbered 0..n so the order is the record."""
    out = sorted(types, key=lambda t: (int(t.get("position") or 0), str(t.get("key") or "")))
    for i, t in enumerate(out):
        t["position"] = i
    return out


def _normalise_request_types(types: list[dict] | None) -> list[dict]:
    out: list[dict] = []
    seen: set[str] = set()
    for i, raw in enumerate(types or []):
        spec = _normalise_request_type(raw, position=i)
        if spec["key"] in seen:
            raise WorkRequestError(f"Request type {spec['key']!r} appears twice")
        seen.add(spec["key"])
        out.append(spec)
    return _renumber(out)


def slugify_key(text: str) -> str:
    """``"Safety PLC"`` → ``safety_plc`` - the key a type gets when the
    manage UI only asks for a label."""
    slug = re.sub(r"[^a-z0-9]+", "_", str(text or "").strip().lower()).strip("_")[:40]
    slug = re.sub(r"^[^a-z]+", "", slug).strip("_")
    return slug


def derive_prefix(key: str, name: str = "") -> str:
    """First three letters of the key (else the name), upper-cased."""
    for source in (key, name):
        letters = re.sub(r"[^A-Za-z]", "", str(source or ""))
        if len(letters) >= 3:
            return letters[:3].upper()
        if letters:
            return letters.upper().ljust(3, "X")
    return "DEP"


async def create_department(session: AsyncSession, data: dict[str, Any]) -> WorkDepartment:
    key = str(data.get("key") or "").strip().lower()
    if not _KEY_RX.match(key):
        raise WorkRequestError("A department key is lowercase letters, digits and underscores (e.g. site_services)")
    name = single_line(clean_text(data.get("name") or ""))[:120]
    if not name:
        raise WorkRequestError("A department needs a name")
    existing = await _departments_by_key(session)
    if key in existing:
        raise ConflictError(f"Department {key!r} already exists")
    prefix = str(data.get("prefix") or "").strip().upper() or derive_prefix(key, name)
    if not re.fullmatch(r"[A-Z]{2,5}", prefix):
        raise WorkRequestError("A reference prefix is 2-5 capital letters (e.g. SVC)")
    if any(d.prefix == prefix for d in existing.values()):
        raise ConflictError(
            f"Reference prefix {prefix!r} is already used by another department - pass a distinct 'prefix'"
        )
    row = WorkDepartment(
        key=key,
        name=name,
        prefix=prefix,
        colour=_colour(data.get("colour")),
        description=clean_text(str(data.get("description") or ""))[:MAX_TEXT_CHARS],
        active=bool(data.get("active", True)),
        position=max((d.position or 0) for d in existing.values()) + 1 if existing else 0,
        lead_user_id=(await _validate_users(session, [data["lead_user_id"]], field="lead_user_id"))[0]
        if data.get("lead_user_id")
        else None,
        member_ids=await _validate_users(session, data.get("member_ids"), field="member_ids"),
        hourly_rate=_money_text(data.get("hourly_rate"), field="hourly_rate"),
        target_days=_target_days(data.get("target_days")),
        stages=_normalise_stages(data.get("stages")),
        request_types=_normalise_request_types(data.get("request_types")),
    )
    session.add(row)
    await session.flush()
    return row


def _colour(value: Any) -> str:
    colour = str(value or "slate").strip().lower()
    if colour not in _COLOURS:
        raise WorkRequestError(f"Unknown colour {colour!r}")
    return colour


async def update_department(session: AsyncSession, dept: WorkDepartment, changes: dict[str, Any]) -> WorkDepartment:
    """Edit a department. The key and the prefix are fixed for life - the
    references already on drawings carry them."""
    if "name" in changes:
        name = single_line(clean_text(changes["name"] or ""))[:120]
        if not name:
            raise WorkRequestError("A department needs a name")
        dept.name = name
    if "colour" in changes:
        dept.colour = _colour(changes["colour"])
    if "description" in changes:
        dept.description = clean_text(str(changes["description"] or ""))[:MAX_TEXT_CHARS]
    if "active" in changes:
        dept.active = bool(changes["active"])
    if "position" in changes and changes["position"] is not None:
        dept.position = max(0, int(changes["position"]))
    if "lead_user_id" in changes:
        lead = changes["lead_user_id"]
        dept.lead_user_id = (await _validate_users(session, [lead], field="lead_user_id"))[0] if lead else None
    if "member_ids" in changes:
        dept.member_ids = await _validate_users(session, changes["member_ids"], field="member_ids")
    if "hourly_rate" in changes:
        dept.hourly_rate = _money_text(changes["hourly_rate"], field="hourly_rate")
    if "target_days" in changes:
        dept.target_days = _target_days(changes["target_days"])
    if "stages" in changes:
        dept.stages = _normalise_stages(changes["stages"])
    if "request_types" in changes:
        dept.request_types = _normalise_request_types(changes["request_types"])
    await session.flush()
    return dept


def stage_spec(dept: WorkDepartment, key: str | None) -> dict | None:
    for s in dept.stages or []:
        if s.get("key") == key:
            return s
    return None


def request_type_spec(dept: WorkDepartment, key: str) -> dict | None:
    """One type by key - RETIRED ones included, so a request raised
    against a since-retired type still resolves its label and fields."""
    for t in dept.request_types or []:
        if t.get("key") == key:
            return t
    return None


def type_keys_of(req: WorkRequest) -> list[str]:
    """Every type on a request, singular column included.

    A row written before ``request_types`` existed (or by a caller that
    still sends the singular) reads back as a one-element list, so no
    caller ever has to know which of the two it came from.
    """
    keys = [str(k or "").strip().lower() for k in (req.request_types or [])]
    keys = [k for k in keys if k]
    if not keys and req.request_type:
        keys = [str(req.request_type).strip().lower()]
    return list(dict.fromkeys(keys))


def normalise_type_keys(request_types: Any, request_type: Any = None) -> list[str]:
    """The submitted list, or the legacy singular. Duplicates collapsed,
    order preserved - the FIRST is the one the ``request_type`` column
    keeps."""
    raw: list[Any]
    if request_types is None:
        raw = [request_type] if request_type else []
    elif isinstance(request_types, (str, bytes)):
        raw = [request_types]
    else:
        raw = list(request_types)
        if not raw and request_type:
            raw = [request_type]
    out: list[str] = []
    for item in raw:
        key = str(item or "").strip().lower()[:60]
        if key and key not in out:
            out.append(key)
    return out


def resolve_request_types(dept: WorkDepartment, keys: list[str]) -> list[dict]:
    """Validate the chosen keys against the department and return their
    specs in the order asked for.

    Every key must be one of the department's types AND still active. A
    department with NO types configured at all takes any key as free text
    - the same escape hatch the single-type path always had, so a fresh
    custom department is usable before anybody fills its catalogue in.
    """
    if not keys:
        raise WorkRequestError("A request type is needed")
    if len(keys) > MAX_REQUEST_TYPES:
        raise WorkRequestError(f"At most {MAX_REQUEST_TYPES} request types on one request")
    if not dept.request_types:
        return []
    out: list[dict] = []
    for key in keys:
        spec = request_type_spec(dept, key)
        if spec is None:
            live = [t["key"] for t in ordered_request_types(dept)]
            raise WorkRequestError(f"Unknown request type {key!r} for {dept.name} (one of {', '.join(live)})")
        if not bool(spec.get("active", True)):
            raise WorkRequestError(
                f"Request type {spec.get('label') or key!r} is retired for {dept.name} - pick a current one"
            )
        out.append(spec)
    return out


def specs_on(dept: WorkDepartment | None, keys: list[str]) -> list[dict]:
    """The specs behind the keys ALREADY on a request - tolerant on
    purpose. A retired type still resolves (it is history now), and a key
    the department no longer declares is simply skipped rather than
    turning every read and every edit of that request into an error."""
    if dept is None:
        return []
    out = []
    for key in keys:
        spec = request_type_spec(dept, key)
        if spec is not None:
            out.append(spec)
    return out


def merged_field_specs(rtypes: list[dict]) -> list[dict]:
    """The UNION of the chosen types' fields, de-duplicated by key, first
    definition wins, in type order - exactly what the raise form should
    render and exactly what the server validates against."""
    out: list[dict] = []
    seen: set[str] = set()
    for rtype in rtypes:
        for spec in rtype.get("fields") or []:
            key = str(spec.get("key") or "")
            if key and key not in seen:
                seen.add(key)
                out.append(dict(spec))
    return out


def merged_checklist_specs(rtypes: list[dict]) -> list[dict]:
    """The UNION of the chosen types' checklists, de-duplicated by key,
    first definition wins, in type order - the same rule ``fields`` uses.

    An item declared ``required`` by ANY of the chosen types is required
    on the request: asking for a switchboard AND its drafting cannot
    quietly drop the drafting sign-off.
    """
    out: list[dict] = []
    at: dict[str, int] = {}
    for rtype in rtypes:
        for spec in rtype.get("checklist") or []:
            key = str(spec.get("key") or "")
            if not key:
                continue
            if key not in at:
                at[key] = len(out)
                out.append({"key": key, "label": spec.get("label") or key, "required": bool(spec.get("required"))})
            elif spec.get("required"):
                out[at[key]]["required"] = True
    return out


def merged_disciplines(rtypes: list[dict]) -> list[str]:
    out: list[str] = []
    for rtype in rtypes:
        for disc in rtype.get("disciplines") or []:
            text = str(disc or "").strip().lower()
            if text and text not in out:
                out.append(text)
    return out


# ── Checklists ───────────────────────────────────────────────────────────


def overrides_of(req: WorkRequest) -> dict[str, Any]:
    """This request's checklist overrides, read defensively.

    A row written before the column existed - or one an older client left
    a half-shape in - reads back as the empty override, which is exactly
    "use the type's list".
    """
    raw = req.checklist_overrides if isinstance(req.checklist_overrides, dict) else {}
    added = [dict(i) for i in (raw.get("added") or []) if isinstance(i, dict) and str(i.get("key") or "")]
    hidden = [str(k) for k in (raw.get("hidden") or []) if str(k or "").strip()]
    edits = {str(k): dict(v) for k, v in (raw.get("edits") or {}).items() if isinstance(v, dict)}
    order = [str(k) for k in (raw.get("order") or []) if str(k or "").strip()]
    return {"added": added, "hidden": hidden, "edits": edits, "order": order}


def derived_checklist_specs(req: WorkRequest, dept: WorkDepartment | None) -> list[dict[str, Any]]:
    """The DEFINITION this request works to: the types' merged checklist
    with this request's overrides applied.

    Each entry carries ``source`` - ``"type"`` for an item inherited from
    the request types, ``"request"`` for one added on this request alone -
    so the UI can show what is inherited and what is a one-off.

    The overrides are a DIFFERENCE, never a copy: an item nobody has
    touched here still reads its label and its ``required`` straight off
    the type, so editing the type still shows through.
    """
    ov = overrides_of(req)
    hidden = set(ov["hidden"])
    out: list[dict[str, Any]] = []
    for spec in merged_checklist_specs(specs_on(dept, type_keys_of(req))):
        if spec["key"] in hidden:
            continue
        out.append({"key": spec["key"], "label": spec["label"], "required": bool(spec["required"]), "source": "type"})

    for item in ov["added"]:
        key = str(item.get("key") or "")
        if any(e["key"] == key for e in out):
            # The TYPE has since declared this key itself. The type's
            # definition governs; the one-off is redundant, not a
            # duplicate row.
            continue
        entry = {
            "key": key,
            "label": str(item.get("label") or key),
            "required": bool(item.get("required")),
            "source": "request",
        }
        anchor = str(item.get("after_key") or "")
        at = next((i for i, e in enumerate(out) if e["key"] == anchor), None)
        if at is None:  # no anchor, or the type has since dropped it
            out.append(entry)
        else:
            out.insert(at + 1, entry)

    for entry in out:
        edit = ov["edits"].get(entry["key"])
        if not isinstance(edit, dict):
            continue
        if edit.get("label"):
            entry["label"] = str(edit["label"])
        if "required" in edit:
            entry["required"] = bool(edit["required"])

    if ov["order"]:
        rank = {key: i for i, key in enumerate(ov["order"])}
        # Stable sort: a key nobody ordered keeps its relative place, at
        # the end, rather than jumping about between two reads.
        out.sort(key=lambda e: (0, rank[e["key"]]) if e["key"] in rank else (1, 0))

    return out[:MAX_CHECKLIST_ITEMS]


def checklist_for(req: WorkRequest, dept: WorkDepartment | None) -> list[dict[str, Any]]:
    """The request's checklist RESOLVED: the DERIVED definition (the
    types' list plus this request's overrides) joined to this request's
    ticks. Ticks for an item the list no longer carries simply stop
    showing - the definition is the list, the state is only ever the
    state, and a hidden item's tick comes back with it.
    """
    state = req.checklist_state if isinstance(req.checklist_state, dict) else {}
    out: list[dict[str, Any]] = []
    for spec in derived_checklist_specs(req, dept):
        tick = state.get(spec["key"])
        tick = tick if isinstance(tick, dict) else {}
        out.append(
            {
                "key": spec["key"],
                "label": spec["label"],
                "required": bool(spec["required"]),
                "source": spec["source"],
                "done": bool(tick.get("done")),
                "by": str(tick.get("by") or "") or None,
                "at": tick.get("at") or None,
            }
        )
    return out


def outstanding_checklist(req: WorkRequest, dept: WorkDepartment | None) -> list[str]:
    """The REQUIRED items still unticked, by label - what a closing move
    is refused with."""
    return [i["label"] for i in checklist_for(req, dept) if i["required"] and not i["done"]]


def _require_checklist_done(req: WorkRequest, dept: WorkDepartment | None) -> None:
    """The gate on completion. Enforced on BOTH paths that complete a
    request - the closing stage and an explicit status change - because a
    rail enforced in one code path is not a rail."""
    outstanding = outstanding_checklist(req, dept)
    if outstanding:
        raise ConflictError(
            f"{req.reference} cannot be completed while these checklist items are outstanding: "
            + ", ".join(outstanding)
        )


async def set_checklist_item(
    session: AsyncSession, req: WorkRequest, key: str, done: bool, *, user_id: str, can_manage: bool
) -> WorkRequest:
    """Tick or untick one item. Department side, like a stage move."""
    dept = await department_or_error(session, req.department)
    _require_department(req, dept, user_id, can_manage)
    if req.status in TERMINAL_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status}")
    wanted = str(key or "").strip().lower()
    items = {i["key"]: i for i in checklist_for(req, dept)}
    spec = items.get(wanted)
    if spec is None:
        raise WorkRequestError(
            f"{req.reference} has no checklist item {key!r}"
            + (f" (one of {', '.join(items)})" if items else " - its request types declare no checklist")
        )
    if bool(spec["done"]) == bool(done):
        return req
    state = dict(req.checklist_state or {})
    state[wanted] = {"done": bool(done), "by": str(user_id), "at": _now().isoformat(timespec="seconds")}
    req.checklist_state = state
    await _log(session, req, user_id=user_id, what="Ticked" if done else "Unticked", detail=spec["label"])
    await session.flush()
    return req


# ── Editing the checklist ON one request ─────────────────────────────────


async def _checklist_edit_context(
    session: AsyncSession, req: WorkRequest, *, user_id: str, can_manage: bool
) -> tuple[WorkDepartment, dict[str, Any], list[dict[str, Any]]]:
    """The three things every checklist edit needs, and the two rails
    every one of them runs into first: who is allowed, and whether this
    request is still open at all."""
    dept = await department_or_error(session, req.department)
    _require_checklist_admin(req, dept, user_id, can_manage)
    if req.status in TERMINAL_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status} - its checklist can no longer be changed")
    return dept, overrides_of(req), checklist_for(req, dept)


def _store_overrides(req: WorkRequest, ov: dict[str, Any]) -> None:
    """Write the override block back, dropping every empty part - a
    reset-by-hand leaves ``{}``, not a shape full of empty lists. A NEW
    dict, so the JSON column is seen as changed."""
    out = {k: v for k, v in ov.items() if v}
    req.checklist_overrides = out


def _checklist_item_or_error(req: WorkRequest, items: list[dict[str, Any]], key: str) -> dict[str, Any]:
    wanted = str(key or "").strip().lower()
    spec = next((i for i in items if i["key"] == wanted), None)
    if spec is None:
        raise WorkRequestError(
            f"{req.reference} has no checklist item {key!r}"
            + (f" (one of {', '.join(i['key'] for i in items)})" if items else " - its checklist is empty")
        )
    return spec


async def add_checklist_item(
    session: AsyncSession,
    req: WorkRequest,
    *,
    label: str,
    required: bool = False,
    after_key: str | None = None,
    user_id: str,
    can_manage: bool,
) -> WorkRequest:
    """Add an item to THIS request only - "Client witness test" on the one
    job that needs it, without touching the type every other job uses.

    The key is slugged from the label, exactly like a type's checklist
    item, and must be free in the request's DERIVED list. Re-adding a key
    that was hidden here brings the inherited item back rather than
    refusing over an item nobody can see.
    """
    _dept, ov, items = await _checklist_edit_context(session, req, user_id=user_id, can_manage=can_manage)
    text = single_line(clean_text(label or ""))[:200]
    if not text:
        raise WorkRequestError("A checklist item needs a label")
    key = slugify_key(text)
    if not _KEY_RX.match(key):
        raise WorkRequestError(f"{text!r} does not make a checklist key - use letters and digits")

    if key in set(ov["hidden"]):
        # It is one of the type's items, hidden here. Bring it back, under
        # the label just asked for.
        ov["hidden"] = [k for k in ov["hidden"] if k != key]
        ov["edits"][key] = {"label": text, "required": bool(required)}
        _store_overrides(req, ov)
        await _log(session, req, user_id=user_id, what="Restored checklist item", detail=text)
        await session.flush()
        return req

    clash = next((i for i in items if i["key"] == key), None)
    if clash is not None:
        raise ConflictError(f"{req.reference} already has a checklist item {key!r} - {clash['label']}")
    if len(items) >= MAX_CHECKLIST_ITEMS:
        raise ConflictError(f"At most {MAX_CHECKLIST_ITEMS} checklist items on one request")

    anchor = str(after_key or "").strip().lower()
    if anchor and not any(i["key"] == anchor for i in items):
        raise WorkRequestError(f"{req.reference} has no checklist item {after_key!r} to add this one after")

    ov["added"] = [*ov["added"], {"key": key, "label": text, "required": bool(required), "after_key": anchor or None}]
    _store_overrides(req, ov)
    await _log(session, req, user_id=user_id, what="Added checklist item", detail=text)
    await session.flush()
    return req


async def update_checklist_item(
    session: AsyncSession,
    req: WorkRequest,
    key: str,
    *,
    label: str | None = None,
    required: bool | None = None,
    user_id: str,
    can_manage: bool,
) -> WorkRequest:
    """Re-word an item, or change whether it gates completion, on THIS
    request. An inherited item is overridden (the type keeps its own
    wording); a request-added one is edited in place. The KEY never
    changes either way - the ticks are keyed by it."""
    _dept, ov, items = await _checklist_edit_context(session, req, user_id=user_id, can_manage=can_manage)
    spec = _checklist_item_or_error(req, items, key)
    wanted = spec["key"]

    text: str | None = None
    if label is not None:
        text = single_line(clean_text(label))[:200]
        if not text:
            raise WorkRequestError("A checklist item needs a label")
    if text is None and required is None:
        return req

    if spec["source"] == "request":
        added = []
        for item in ov["added"]:
            if str(item.get("key") or "") == wanted:
                item = dict(item)
                if text is not None:
                    item["label"] = text
                if required is not None:
                    item["required"] = bool(required)
            added.append(item)
        ov["added"] = added
    else:
        edit = dict(ov["edits"].get(wanted) or {})
        if text is not None:
            edit["label"] = text
        if required is not None:
            edit["required"] = bool(required)
        ov["edits"][wanted] = edit

    _store_overrides(req, ov)
    await _log(session, req, user_id=user_id, what="Changed checklist item", detail=text or spec["label"])
    await session.flush()
    return req


async def remove_checklist_item(
    session: AsyncSession, req: WorkRequest, key: str, *, user_id: str, can_manage: bool
) -> WorkRequest:
    """Drop an item from THIS request.

    One added here is removed outright. One inherited from the type is
    HIDDEN, never deleted: the type still declares it, and a later change
    of request type must be able to bring it back. A TICKED item is
    refused - a list is not tidied by deleting the evidence.
    """
    _dept, ov, items = await _checklist_edit_context(session, req, user_id=user_id, can_manage=can_manage)
    spec = _checklist_item_or_error(req, items, key)
    wanted = spec["key"]
    if spec["done"]:
        raise ConflictError(f"{spec['label']!r} is ticked on {req.reference} - untick it first, then remove it")

    if spec["source"] == "request":
        ov["added"] = [i for i in ov["added"] if str(i.get("key") or "") != wanted]
        ov["edits"].pop(wanted, None)
        what = "Removed checklist item"
    else:
        if wanted not in ov["hidden"]:
            ov["hidden"] = [*ov["hidden"], wanted]
        what = "Hid checklist item"
    ov["order"] = [k for k in ov["order"] if k != wanted]
    # Anything anchored to it falls back to the end of the list.
    ov["added"] = [dict(i, after_key=None) if str(i.get("after_key") or "") == wanted else i for i in ov["added"]]
    _store_overrides(req, ov)
    await _log(session, req, user_id=user_id, what=what, detail=spec["label"])
    await session.flush()
    return req


async def reorder_checklist(
    session: AsyncSession, req: WorkRequest, keys: list[str], *, user_id: str, can_manage: bool
) -> WorkRequest:
    """``{keys: [...]}`` - the order this request's list reads in. A key
    left out keeps its relative order, at the end, and a key that is not
    on the list is ignored rather than refused: the UI sends what it has
    on screen, and the type may have moved underneath it."""
    _dept, ov, items = await _checklist_edit_context(session, req, user_id=user_id, can_manage=can_manage)
    live = {i["key"] for i in items}
    wanted: list[str] = []
    for raw in keys or []:
        k = str(raw or "").strip().lower()
        if k in live and k not in wanted:
            wanted.append(k)
    ov["order"] = wanted
    _store_overrides(req, ov)
    await _log(session, req, user_id=user_id, what="Reordered checklist", detail=", ".join(wanted))
    await session.flush()
    return req


async def reset_checklist(session: AsyncSession, req: WorkRequest, *, user_id: str, can_manage: bool) -> WorkRequest:
    """Drop EVERY override and go back to what the request types declare.

    The ticks are left alone: an item that still exists keeps the tick it
    had, and one that only existed here simply stops showing - the same
    rule a retired type's item already follows.
    """
    await _checklist_edit_context(session, req, user_id=user_id, can_manage=can_manage)
    if not (req.checklist_overrides or {}):
        return req
    req.checklist_overrides = {}
    await _log(session, req, user_id=user_id, what="Reset checklist", detail="back to the request type's list")
    await session.flush()
    return req


# ── Turnaround targets ───────────────────────────────────────────────────


def add_working_days(start: date, days: int) -> date:
    """``start`` plus ``days`` WORKING days - Saturday and Sunday skipped.

    A three-day turnaround accepted on a Friday is due Wednesday, not
    Monday: the department does not work the weekend and must not be
    judged as though it did. Public holidays are not modelled (they are
    a per-site calendar this module does not own).
    """
    day = start
    remaining = max(0, int(days))
    while remaining > 0:
        day += timedelta(days=1)
        if day.weekday() < 5:
            remaining -= 1
    return day


def target_date_for(req: WorkRequest, dept: WorkDepartment | None) -> str | None:
    """When this request is due by the department's own turnaround target.

    ``None`` until BOTH a target and an acceptance day exist - a request
    nobody has accepted yet cannot be late.
    """
    if dept is None or dept.target_days is None or not req.accepted_at:
        return None
    try:
        accepted = date.fromisoformat(req.accepted_at)
    except ValueError:
        return None
    return add_working_days(accepted, int(dept.target_days)).isoformat()


def lateness(
    req: WorkRequest, dept: WorkDepartment | None, *, today: date | None = None
) -> tuple[str | None, int | None, bool]:
    """``(target_date, days_late, is_late)``.

    ``days_late`` is measured to today while the request is open, and to
    the day it stopped being open once it is not - a finished request's
    lateness is a fact about when it finished, not about how long ago.
    """
    target = target_date_for(req, dept)
    if target is None:
        return None, None, False
    now = today or _today()
    if req.status in ACTIVE_STATUSES:
        reference_day = now
    elif req.closed_at is not None:
        reference_day = req.closed_at.date()
    else:
        reference_day = now
    days_late = max(0, (reference_day - date.fromisoformat(target)).days)
    return target, days_late, bool(days_late > 0 and req.status in ACTIVE_STATUSES)


# ── The department's own type catalogue (manage) ─────────────────────────


async def requests_using_type(session: AsyncSession, department: str, type_key: str) -> int:
    """How many requests name this type - singular column or list."""
    pattern = "%" + _like_escape(f'"{type_key}"') + "%"
    return int(
        await session.scalar(
            select(func.count())
            .select_from(WorkRequest)
            .where(
                WorkRequest.department == department,
                or_(
                    WorkRequest.request_type == type_key,
                    cast(WorkRequest.request_types, String).like(pattern, escape="\\"),
                ),
            )
        )
        or 0
    )


async def create_request_type(session: AsyncSession, dept: WorkDepartment, data: dict[str, Any]) -> dict:
    """Add one type to a department. The key is slugged from the label
    when it is not given, and is unique within the department."""
    label = single_line(clean_text(data.get("label") or ""))[:200]
    key = str(data.get("key") or "").strip().lower() or slugify_key(label)
    if not key:
        raise WorkRequestError("A request type needs a key or a label to make one from")
    if not label:
        raise WorkRequestError("A request type needs a label")
    existing = ordered_request_types(dept, include_inactive=True)
    if any(t["key"] == key for t in existing):
        raise ConflictError(f"{dept.name} already has a request type {key!r}")
    position = data.get("position")
    spec = _normalise_request_type(
        {**data, "key": key, "label": label},
        position=int(position) if position is not None else len(existing),
    )
    dept.request_types = _renumber([*existing, spec])
    await session.flush()
    return next(t for t in dept.request_types if t["key"] == key)


async def update_request_type(
    session: AsyncSession, dept: WorkDepartment, type_key: str, changes: dict[str, Any]
) -> dict:
    """Edit one type in place. The key is fixed - the requests already
    raised against it carry it."""
    key = str(type_key or "").strip().lower()
    existing = ordered_request_types(dept, include_inactive=True)
    current = next((t for t in existing if t["key"] == key), None)
    if current is None:
        raise NotFoundError(f"{dept.name} has no request type {key!r}")
    merged = {**current, **{k: v for k, v in changes.items() if k != "key"}, "key": key}
    spec = _normalise_request_type(merged, position=current["position"])
    dept.request_types = _renumber([spec if t["key"] == key else t for t in existing])
    await session.flush()
    return next(t for t in dept.request_types if t["key"] == key)


async def delete_request_type(session: AsyncSession, dept: WorkDepartment, type_key: str) -> None:
    """Remove a type NOTHING has ever been raised against.

    A type in use is never deleted - the requests carrying it would lose
    their label and their fields. Retire it instead (``active: false``):
    it stops being offered and stays readable where it was used.
    """
    key = str(type_key or "").strip().lower()
    existing = ordered_request_types(dept, include_inactive=True)
    if not any(t["key"] == key for t in existing):
        raise NotFoundError(f"{dept.name} has no request type {key!r}")
    used = await requests_using_type(session, dept.key, key)
    if used:
        raise ConflictError(
            f"{used} request{'s' if used != 1 else ''} already use {key!r} - "
            f"set it inactive to retire it instead of deleting it"
        )
    dept.request_types = _renumber([t for t in existing if t["key"] != key])
    await session.flush()


async def reorder_request_types(session: AsyncSession, dept: WorkDepartment, keys: list[str]) -> list[dict]:
    """``{keys: [...]}`` sets the order. Every key must be one of this
    department's; anything left out keeps its relative order at the end,
    so a stale UI list can never drop a type."""
    existing = ordered_request_types(dept, include_inactive=True)
    known = {t["key"] for t in existing}
    wanted: list[str] = []
    for raw in keys or []:
        key = str(raw or "").strip().lower()
        if key not in known:
            raise WorkRequestError(f"{dept.name} has no request type {key!r}")
        if key not in wanted:
            wanted.append(key)
    order = {key: i for i, key in enumerate(wanted)}
    tail = len(wanted)
    for t in existing:
        t["position"] = order.get(t["key"], tail + t["position"])
    dept.request_types = _renumber(existing)
    await session.flush()
    return dept.request_types


# ── Who may do what ──────────────────────────────────────────────────────


def is_department_side(dept: WorkDepartment | None, user_id: str) -> bool:
    """Is this person the department, for the purposes of the rails?

    The lead and the members are. A department with NEITHER configured
    yet is open to everyone with the update permission - a fresh install
    must not lock every board until somebody fills in the roster.
    """
    if dept is None:
        return False
    people = {str(m) for m in (dept.member_ids or [])}
    if dept.lead_user_id:
        people.add(str(dept.lead_user_id))
    if not people:
        return True
    return str(user_id) in people


def may_act_for_department(req: WorkRequest, dept: WorkDepartment | None, user_id: str, can_manage: bool) -> bool:
    uid = str(user_id)
    return (
        can_manage
        or is_department_side(dept, uid)
        or uid in {str(a) for a in (req.assignee_ids or [])}
        or uid == str(req.responsible_user_id or "")
    )


def may_update(req: WorkRequest, dept: WorkDepartment | None, user_id: str, can_manage: bool) -> bool:
    return may_act_for_department(req, dept, user_id, can_manage) or str(user_id) == str(req.raised_by_id or "")


def _require_update(req: WorkRequest, dept: WorkDepartment | None, user_id: str, can_manage: bool) -> None:
    if not may_update(req, dept, user_id, can_manage):
        raise NotPermitted(
            f"Only the requester, the {dept.name if dept else req.department} team or a manager can change "
            f"{req.reference}"
        )


def _require_department(req: WorkRequest, dept: WorkDepartment | None, user_id: str, can_manage: bool) -> None:
    if not may_act_for_department(req, dept, user_id, can_manage):
        raise NotPermitted(
            f"Only the {dept.name if dept else req.department} team (or a manager) can do that on {req.reference}"
        )


def may_edit_checklist(dept: WorkDepartment | None, user_id: str, can_manage: bool) -> bool:
    """Who may change the SHAPE of a request's checklist.

    Deliberately narrower than ticking: a fitter says an item is done, but
    what the list IS belongs to the department's lead or a manager.
    A department with no lead configured is manager-only - the opposite of
    the ``is_department_side`` escape hatch, because a nobody-configured
    roster must not hand the list to the whole install.
    """
    if can_manage:
        return True
    return dept is not None and bool(dept.lead_user_id) and str(user_id) == str(dept.lead_user_id)


def _require_checklist_admin(req: WorkRequest, dept: WorkDepartment | None, user_id: str, can_manage: bool) -> None:
    if not may_edit_checklist(dept, user_id, can_manage):
        raise NotPermitted(
            f"Only the {dept.name if dept else req.department} lead or a manager can change this checklist."
        )


# ── References ───────────────────────────────────────────────────────────


def format_reference(prefix: str, number: int) -> str:
    """``WR-ENG-000001``. Widens past six digits rather than wrapping."""
    return f"{REFERENCE_HOUSE}-{prefix}-{number:0{REFERENCE_DIGITS}d}"


async def _highest_existing(session: AsyncSession, prefix: str) -> int:
    """The biggest number already issued on this prefix - seeds the counter
    on first use so a re-install never re-issues a number on a drawing."""
    pattern = f"{REFERENCE_HOUSE}-{prefix}-%"
    rows = (
        (await session.execute(select(WorkRequest.reference).where(WorkRequest.reference.like(pattern))))
        .scalars()
        .all()
    )
    rx = re.compile(rf"^{re.escape(REFERENCE_HOUSE)}-{re.escape(prefix)}-(\d+)$")
    highest = 0
    for ref in rows:
        m = rx.match(str(ref or ""))
        if m:
            highest = max(highest, int(m.group(1)))
    return highest


async def peek_reference(session: AsyncSession, dept: WorkDepartment) -> str:
    """The NEXT reference, without burning it - for a raise-form preview."""
    row = (
        await session.execute(select(WorkRequestCounter).where(WorkRequestCounter.prefix == dept.prefix))
    ).scalar_one_or_none()
    current = row.value if row is not None else await _highest_existing(session, dept.prefix)
    return format_reference(dept.prefix, int(current or 0) + 1)


async def next_reference(session: AsyncSession, dept: WorkDepartment) -> str:
    """Burn and return the next reference for this department. Never re-issued.

    The counter row is taken WITH A LOCK: two people raising on the same
    department in the same second must not both read the same number. The
    FIRST mint has its own race - both find no row, both insert - so the
    insert runs in a savepoint and the loser goes back around to lock the
    winner's row instead of surfacing an IntegrityError as a 500.
    """
    from sqlalchemy.exc import IntegrityError

    prefix = dept.prefix
    row = None
    for attempt in (1, 2):
        row = (
            await session.execute(
                select(WorkRequestCounter).where(WorkRequestCounter.prefix == prefix).with_for_update()
            )
        ).scalar_one_or_none()
        if row is not None:
            break
        try:
            async with session.begin_nested():
                row = WorkRequestCounter(prefix=prefix, value=await _highest_existing(session, prefix))
                session.add(row)
                await session.flush()
            break
        except IntegrityError as exc:
            row = None
            if attempt == 2:
                raise WorkRequestError(f"The reference counter for {prefix} could not be created - try again") from exc
    assert row is not None  # both loop exits above set it
    row.value = int(row.value or 0) + 1
    await session.flush()
    return format_reference(prefix, row.value)


# ── Field validation ─────────────────────────────────────────────────────


def _validate_fields(field_specs: list[dict] | None, fields: dict[str, Any] | None, *, strict: bool) -> dict[str, Any]:
    """The chosen types' questions - the UNION of their fields - typed and
    (when ``strict``) complete. Keys no type declares are kept as free
    text so a department can add a column before it adds it to a type."""
    raw = clean_text(dict(fields or {}))
    specs = {f["key"]: f for f in (field_specs or [])}
    out: dict[str, Any] = {}
    for key, value in raw.items():
        key = str(key)[:60]
        spec = specs.get(key)
        if spec is None:
            if isinstance(value, (str, int, float, bool)) or value is None:
                out[key] = value if not isinstance(value, str) else value[:MAX_TEXT_CHARS]
            continue
        ftype = spec.get("type", "text")
        label = spec.get("label") or key
        if value is None or value == "":
            out[key] = None
            continue
        if ftype in ("text", "area"):
            out[key] = str(value)[: (MAX_TITLE_CHARS if ftype == "text" else MAX_TEXT_CHARS)]
        elif ftype == "date":
            out[key] = _iso(value, field=label)
        elif ftype == "number":
            try:
                out[key] = float(value)
            except (TypeError, ValueError):
                raise WorkRequestError(f"{label} must be a number, not {value!r}") from None
        elif ftype == "bool":
            if isinstance(value, str):
                out[key] = value.strip().lower() in ("1", "true", "yes", "y", "on")
            else:
                out[key] = bool(value)
        elif ftype == "select":
            text = str(value).strip()
            if text not in (spec.get("options") or []):
                raise WorkRequestError(f"{label} must be one of: {', '.join(spec.get('options') or [])}")
            out[key] = text
        elif ftype == "url":
            text = str(value).strip()[:2000]
            if not re.match(r"^https?://", text, re.I):
                raise WorkRequestError(f"{label} must be a web link starting with http:// or https://")
            out[key] = text
    if strict:
        missing = [
            spec.get("label") or key
            for key, spec in specs.items()
            if spec.get("required") and (out.get(key) is None or out.get(key) == "")
        ]
        if missing:
            raise WorkRequestError("These are needed before it can be raised: " + ", ".join(missing))
    return out


def _validate_links(links: list[Any] | None) -> list[dict]:
    out: list[dict] = []
    for raw in links or []:
        if isinstance(raw, dict):
            label = single_line(clean_text(raw.get("label") or ""))[:200]
            url = str(raw.get("url") or "").strip()[:2000]
        else:
            label, url = "", str(raw or "").strip()[:2000]
        if not url:
            continue
        if not re.match(r"^https?://", url, re.I):
            raise WorkRequestError(f"Link {label or url!r} must start with http:// or https://")
        out.append({"label": label or url, "url": url})
    if len(out) > MAX_LINKS:
        raise WorkRequestError(f"At most {MAX_LINKS} links")
    return out


def _validate_disciplined(value: dict | None, disciplines: list[str] | None, *, field: str, numeric: bool) -> dict:
    allowed = set(disciplines or [])
    out: dict = {}
    for key, raw in (value or {}).items():
        disc = str(key or "").strip().lower()[:40]
        if not disc:
            continue
        if allowed and disc not in allowed:
            raise WorkRequestError(
                f"{field}: {disc!r} is not a discipline of this request type ({', '.join(sorted(allowed))})"
            )
        if numeric:
            hours = _hours(raw, field=f"{field}[{disc}]")
            if hours is not None:
                out[disc] = hours
        else:
            text = single_line(clean_text(raw))[:60]
            if text:
                out[disc] = text
    return out


def _priority(value: Any) -> str:
    p = str(value or "normal").strip().lower()
    if p not in PRIORITIES:
        raise WorkRequestError(f"Priority must be one of {', '.join(PRIORITIES)}")
    return p


async def _validate_depends_on(session: AsyncSession, ids: list[str] | None, *, self_id: uuid.UUID | None) -> list[str]:
    seen: list[str] = []
    for raw in ids or []:
        text = str(raw or "").strip()
        if text and text not in seen:
            seen.append(text)
    if not seen:
        return []
    as_uuid = [_as_uuid(s) for s in seen]
    if any(u is None for u in as_uuid):
        raise WorkRequestError("depends_on_ids must be request ids")
    if self_id is not None and str(self_id) in seen:
        raise WorkRequestError("A request cannot depend on itself")
    found = {
        str(r) for r in (await session.execute(select(WorkRequest.id).where(WorkRequest.id.in_(as_uuid)))).scalars()
    }
    missing = [s for s in seen if s not in found]
    if missing:
        raise WorkRequestError(f"depends_on_ids: unknown request {missing[0]}")
    return seen


async def _validate_parent(session: AsyncSession, parent_id: Any, *, self_id: uuid.UUID | None) -> uuid.UUID | None:
    if not parent_id:
        return None
    pid = _as_uuid(parent_id)
    if pid is None:
        raise WorkRequestError("parent_id must be a request id")
    if self_id is not None and pid == self_id:
        raise WorkRequestError("A request cannot be its own parent")
    if await session.get(WorkRequest, pid) is None:
        raise WorkRequestError(f"parent_id: unknown request {parent_id}")
    return pid


# ── The programme: the activity this feeds, the lines it draws on ────────


async def _validate_schedule_activity(session: AsyncSession, value: Any, *, project_id: uuid.UUID) -> uuid.UUID | None:
    """One programme activity, and it must be on THIS job.

    Read straight off the schedule tables - no service import, so the
    programme module owns its own rules and this one only ever looks.
    """
    if value in (None, "", []):
        return None
    aid = _as_uuid(value)
    if aid is None:
        raise WorkRequestError(f"schedule_activity_id must be an activity id, not {value!r}")
    try:
        from app.modules.schedule.models import Activity, Schedule
    except ImportError:  # pragma: no cover - schedule module absent
        raise WorkRequestError("The programme module is not installed on this deployment") from None
    owner = (
        await session.execute(
            select(Schedule.project_id).join(Activity, Activity.schedule_id == Schedule.id).where(Activity.id == aid)
        )
    ).scalar_one_or_none()
    if owner is None:
        raise NotFoundError(f"Unknown programme activity {value}")
    if str(owner) != str(project_id):
        raise WorkRequestError(f"Programme activity {value} is on another job - a request only points at its own")
    return aid


async def _validate_boq_positions(session: AsyncSession, values: Any, *, project_id: uuid.UUID) -> list[str]:
    """The estimate lines this work draws on, all on THIS job."""
    wanted: list[str] = []
    for raw in values or []:
        text = str(raw or "").strip()
        if text and text not in wanted:
            wanted.append(text)
    if not wanted:
        return []
    if len(wanted) > MAX_BOQ_POSITIONS:
        raise WorkRequestError(f"At most {MAX_BOQ_POSITIONS} estimate lines on one request")
    as_uuid = [_as_uuid(w) for w in wanted]
    if any(u is None for u in as_uuid):
        raise WorkRequestError("boq_position_ids: every entry must be an estimate line id")
    try:
        from app.modules.boq.models import BOQ, Position
    except ImportError:  # pragma: no cover - boq module absent
        raise WorkRequestError("The estimating module is not installed on this deployment") from None
    rows = (
        await session.execute(
            select(Position.id, BOQ.project_id).join(BOQ, Position.boq_id == BOQ.id).where(Position.id.in_(as_uuid))
        )
    ).all()
    owners = {str(pid): str(proj) for pid, proj in rows}
    missing = [w for w in wanted if w not in owners]
    if missing:
        raise NotFoundError(f"Unknown estimate line {missing[0]}")
    elsewhere = [w for w in wanted if owners[w] != str(project_id)]
    if elsewhere:
        raise WorkRequestError(f"Estimate line {elsewhere[0]} is on another job - a request only points at its own")
    return wanted


# ── Activity ─────────────────────────────────────────────────────────────


async def _log(
    session: AsyncSession,
    req: WorkRequest,
    *,
    user_id: str,
    user_name: str | None = None,
    what: str,
    detail: str = "",
    kind: str = "system",
    mention_ids: list[str] | None = None,
) -> WorkRequestComment:
    if kind not in COMMENT_KINDS:
        raise WorkRequestError(f"Unknown comment kind {kind!r}")
    row = WorkRequestComment(
        request_id=req.id,
        author_id=str(user_id or ""),
        author_name=user_name if user_name is not None else await _user_name(session, user_id),
        body=clean_text(str(what or ""))[:MAX_COMMENT_CHARS],
        detail=clean_text(str(detail or ""))[:MAX_TEXT_CHARS],
        kind=kind,
        mention_ids=list(mention_ids or []),
    )
    session.add(row)
    await session.flush()
    return row


# ── Raising ──────────────────────────────────────────────────────────────


async def create_request(
    session: AsyncSession,
    *,
    project_id: uuid.UUID,
    department: str,
    request_type: str = "",
    request_types: list[str] | None = None,
    title: str,
    description: str = "",
    cost_centres: dict | None = None,
    estimated_hours: dict | None = None,
    quoted_hours: Any = None,
    info_required_by: Any = None,
    due_date: Any = None,
    priority: str = "normal",
    links: list | None = None,
    fields: dict | None = None,
    assignee_ids: list[str] | None = None,
    responsible_user_id: str | None = None,
    depends_on_ids: list[str] | None = None,
    parent_id: Any = None,
    draft: bool = False,
    user_id: str,
    notify: bool = True,
) -> WorkRequest:
    """Raise a request. Mints the reference, lays it in the department's
    first stage, writes the first activity line and tells the department."""
    from sqlalchemy.exc import IntegrityError

    await ensure_seeded(session)
    dept = await department_or_error(session, department)
    if not dept.active:
        raise WorkRequestError(f"{dept.name} is not taking requests (inactive)")
    type_keys = normalise_type_keys(request_types, request_type)
    rtypes = resolve_request_types(dept, type_keys)
    field_specs = merged_field_specs(rtypes)
    disciplines = merged_disciplines(rtypes)
    title = single_line(clean_text(title))[:MAX_TITLE_CHARS]
    if not title:
        raise WorkRequestError("A request needs a title")

    raised_by_name = await _user_name(session, user_id)
    clean_fields = _validate_fields(field_specs, fields, strict=not draft)
    assignees = await _validate_users(session, assignee_ids, field="assignee_ids")
    responsible = (
        (await _validate_users(session, [responsible_user_id], field="responsible_user_id"))[0]
        if responsible_user_id
        else None
    )
    depends = await _validate_depends_on(session, depends_on_ids, self_id=None)
    parent = await _validate_parent(session, parent_id, self_id=None)
    first_stage = (dept.stages or [{}])[0].get("key") if dept.stages else None
    now = _now().isoformat(timespec="seconds")

    for _attempt in range(25):
        reference = await next_reference(session, dept)
        req = WorkRequest(
            project_id=project_id,
            department=dept.key,
            request_type=type_keys[0],
            request_types=list(type_keys),
            reference=reference,
            title=title,
            description=clean_text(str(description or ""))[:MAX_TEXT_CHARS],
            status="draft" if draft else "submitted",
            stage=first_stage,
            stage_history=(
                [{"stage": first_stage, "at": now, "by_id": str(user_id), "by_name": raised_by_name, "note": "Raised"}]
                if first_stage
                else []
            ),
            raised_by_id=str(user_id),
            raised_by_name=raised_by_name,
            assignee_ids=assignees,
            responsible_user_id=responsible,
            cost_centres=_validate_disciplined(cost_centres, disciplines, field="cost_centres", numeric=False),
            estimated_hours=_validate_disciplined(estimated_hours, disciplines, field="estimated_hours", numeric=True),
            quoted_hours=_hours(quoted_hours, field="quoted_hours"),
            info_required_by=_iso(info_required_by, field="info_required_by"),
            due_date=_iso(due_date, field="due_date"),
            priority=_priority(priority),
            links=_validate_links(links),
            fields=clean_fields,
            planner_uploaded=bool(clean_fields.get("planner_uploaded") or False),
            ball_in_court="requester" if draft else "department",
            depends_on_ids=depends,
            parent_id=parent,
        )
        try:
            async with session.begin_nested():
                session.add(req)
                await session.flush()
            break
        except IntegrityError:
            # The counter was seeded below a reference already on the
            # table (a hand-restored row); go around and mint the next.
            logger.warning("work_requests: reference %s collided, re-minting", reference)
            continue
    else:
        raise WorkRequestError("Could not allocate a reference - try again")

    await _log(
        session,
        req,
        user_id=user_id,
        user_name=raised_by_name,
        what="Raised" if not draft else "Saved as draft",
        detail=f"{req.reference} for {dept.name}",
    )
    if notify and not draft:
        from app.modules.work_requests import events, notifying

        await notifying.request_raised(session, req, dept, actor_id=str(user_id))
        # People put on it at raise time hear too - once: a member who is
        # also an assignee already got the department's bell.
        dept_people = set(notifying.department_people(dept))
        on_it = [*assignees, *([responsible] if responsible else [])]
        await notifying.assigned(
            session, req, [a for a in dict.fromkeys(on_it) if a not in dept_people], actor_id=str(user_id)
        )
        events.request_raised(req)
    return req


# ── Editing ──────────────────────────────────────────────────────────────


def allowed_transitions(req: WorkRequest) -> list[str]:
    return list(TRANSITIONS.get(req.status, ()))


async def _set_status(
    session: AsyncSession,
    req: WorkRequest,
    new_status: str,
    *,
    user_id: str,
    can_manage: bool,
    note: str = "",
    force: bool = False,
    dept: WorkDepartment | None = None,
) -> None:
    """The ONE path a status changes through. Refuses an illegal move
    with what is allowed, and enforces who may close.

    ``force`` is for the board: a closing stage completes from ANY open
    state and a first stage move starts the work from ``submitted`` -
    those skip the table but never leave a terminal state.
    """
    new_status = str(new_status or "").strip().lower()
    if new_status not in STATUSES:
        raise WorkRequestError(f"Unknown status {new_status!r} (one of {', '.join(STATUSES)})")
    if new_status == req.status:
        return
    allowed = allowed_transitions(req)
    if force and req.status not in TERMINAL_STATUSES:
        allowed = [new_status]
    if new_status not in allowed:
        raise TransitionError(
            f"{req.reference} is {req.status}; it cannot move to {new_status}"
            + (f" (allowed: {', '.join(allowed)})" if allowed else " - it is final"),
            allowed,
        )
    if new_status == "closed" and not (can_manage or str(user_id) == str(req.raised_by_id)):
        raise NotPermitted(
            f"Only the requester ({req.raised_by_name or 'who raised it'}) or a manager closes a request"
        )
    if new_status == "complete":
        _require_checklist_done(req, dept if dept is not None else await department_or_error(session, req.department))
    previous = req.status
    req.status = new_status
    # The turnaround clock starts the first time the department takes it
    # on, whichever way that happened - accepting it, or simply moving it
    # off the intake stage. Written once: a reopen does not restart it.
    if req.accepted_at is None and new_status in ("accepted", "in_progress", "review", "on_hold", "complete"):
        req.accepted_at = _today().isoformat()
    if new_status in ("complete",):
        req.ball_in_court = "requester"
    elif new_status in ("submitted", "accepted", "in_progress", "review"):
        if req.needs_info is None:
            req.ball_in_court = "department"
    if new_status in TERMINAL_STATUSES:
        req.closed_at = _now()
    elif previous in TERMINAL_STATUSES:
        req.closed_at = None
    await _log(session, req, user_id=user_id, what=f"Status {previous} → {new_status}", detail=single_line(note)[:2000])
    from app.modules.work_requests import events, notifying

    await notifying.status_changed(session, req, previous, actor_id=str(user_id))
    events.status_changed(req, previous)


async def update_request(
    session: AsyncSession,
    req: WorkRequest,
    changes: dict[str, Any],
    *,
    user_id: str,
    can_manage: bool,
) -> WorkRequest:
    """PATCH. Only keys present in ``changes`` are touched."""
    dept = await department_or_error(session, req.department)
    _require_update(req, dept, user_id, can_manage)
    if req.status in TERMINAL_STATUSES and set(changes) - {"status"}:
        raise ConflictError(f"{req.reference} is {req.status} and cannot be edited")
    touched: list[str] = []

    # The types come FIRST: the field union and the allowed disciplines
    # every other key is judged against are derived from them, so a PATCH
    # that swaps the types and fills the new type's fields in one call
    # must be validated against the new set, not the old.
    if "request_types" in changes or "request_type" in changes:
        keys = normalise_type_keys(changes.get("request_types"), changes.get("request_type"))
        resolve_request_types(dept, keys)
        if keys != type_keys_of(req):
            req.request_types = list(keys)
            req.request_type = keys[0]
            touched.append("request type" if len(keys) == 1 else "request types")
    rtypes = specs_on(dept, type_keys_of(req))
    field_specs = merged_field_specs(rtypes)
    disciplines = merged_disciplines(rtypes)

    if "title" in changes:
        title = single_line(clean_text(changes["title"]))[:MAX_TITLE_CHARS]
        if not title:
            raise WorkRequestError("A request needs a title")
        req.title = title
        touched.append("title")
    if "description" in changes:
        req.description = clean_text(str(changes["description"] or ""))[:MAX_TEXT_CHARS]
        touched.append("description")
    if "cost_centres" in changes:
        req.cost_centres = _validate_disciplined(
            changes["cost_centres"], disciplines, field="cost_centres", numeric=False
        )
        touched.append("cost centres")
    if "estimated_hours" in changes:
        req.estimated_hours = _validate_disciplined(
            changes["estimated_hours"], disciplines, field="estimated_hours", numeric=True
        )
        touched.append("estimated hours")
    if "quoted_hours" in changes:
        req.quoted_hours = _hours(changes["quoted_hours"], field="quoted_hours")
        touched.append("quoted hours")
    if "hours_to_complete" in changes:
        req.hours_to_complete = _hours(changes["hours_to_complete"], field="hours_to_complete")
        touched.append("hours to complete")
    for key, label in (
        ("info_required_by", "info required by"),
        ("due_date", "due date"),
        ("scheduled_start", "scheduled start"),
        ("scheduled_end", "scheduled end"),
        ("delivered_at", "delivered"),
        ("tested_at", "tested"),
    ):
        if key in changes:
            setattr(req, key, _iso(changes[key], field=key))
            touched.append(label)
    if "priority" in changes:
        req.priority = _priority(changes["priority"])
        touched.append("priority")
    if "links" in changes:
        req.links = _validate_links(changes["links"])
        touched.append("links")
    if "fields" in changes:
        req.fields = _validate_fields(field_specs, changes["fields"], strict=False)
        if "planner_uploaded" in req.fields and "planner_uploaded" not in changes:
            req.planner_uploaded = bool(req.fields.get("planner_uploaded") or False)
        touched.append("fields")
    if "planner_uploaded" in changes:
        req.planner_uploaded = bool(changes["planner_uploaded"])
        if "planner_uploaded" in (req.fields or {}):
            req.fields = {**req.fields, "planner_uploaded": req.planner_uploaded}
        touched.append("planner uploaded" if req.planner_uploaded else "planner not uploaded")
    if "assignee_ids" in changes or "responsible_user_id" in changes:
        await assign(
            session,
            req,
            assignee_ids=changes.get("assignee_ids", req.assignee_ids),
            responsible_user_id=changes.get("responsible_user_id", req.responsible_user_id),
            user_id=user_id,
            can_manage=can_manage,
            _checked=True,
        )
    if "depends_on_ids" in changes:
        req.depends_on_ids = await _validate_depends_on(session, changes["depends_on_ids"], self_id=req.id)
        touched.append("dependencies")
    if "parent_id" in changes:
        req.parent_id = await _validate_parent(session, changes["parent_id"], self_id=req.id)
        touched.append("parent")
    if "schedule_activity_id" in changes:
        req.schedule_activity_id = await _validate_schedule_activity(
            session, changes["schedule_activity_id"], project_id=req.project_id
        )
        touched.append("programme activity")
    if "boq_position_ids" in changes:
        req.boq_position_ids = await _validate_boq_positions(
            session, changes["boq_position_ids"], project_id=req.project_id
        )
        touched.append("estimate lines")
    if "is_template" in changes:
        req.is_template = bool(changes["is_template"])
        touched.append("marked as a template" if req.is_template else "no longer a template")

    if touched:
        await _log(session, req, user_id=user_id, what="Edited", detail=", ".join(touched))

    if "stage" in changes and changes["stage"] is not None and changes["stage"] != req.stage:
        await move_stage(
            session,
            req,
            changes["stage"],
            note=str(changes.get("stage_note") or ""),
            user_id=user_id,
            can_manage=can_manage,
            _checked=True,
        )
    if "status" in changes and changes["status"]:
        await _set_status(
            session,
            req,
            changes["status"],
            user_id=user_id,
            can_manage=can_manage,
            note=str(changes.get("status_note") or ""),
            dept=dept,
        )
    await session.flush()
    return req


async def move_stage(
    session: AsyncSession,
    req: WorkRequest,
    stage: str,
    *,
    note: str = "",
    user_id: str,
    can_manage: bool,
    _checked: bool = False,
) -> WorkRequest:
    """Move to a stage on the department's board. Appends to the history;
    the first move after acceptance starts the work; a closing stage
    completes it."""
    dept = await department_or_error(session, req.department)
    if not _checked:
        _require_department(req, dept, user_id, can_manage)
    key = str(stage or "").strip().lower()
    spec = stage_spec(dept, key)
    if spec is None:
        raise WorkRequestError(
            f"Unknown stage {stage!r} for {dept.name} (one of {', '.join(s['key'] for s in dept.stages or [])})"
        )
    if req.status in TERMINAL_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status} - reopen it before moving it")
    if key == req.stage:
        raise ConflictError(f"{req.reference} is already at {spec['name']}")
    # BEFORE anything moves: a closing stage completes the request, and a
    # required checklist item still unticked stops it - the board must not
    # end up on the closing column with the work refused behind it.
    if spec.get("closes") and req.status != "complete":
        _require_checklist_done(req, dept)
    previous = req.stage
    req.stage = key
    req.stage_history = [
        *list(req.stage_history or []),
        {
            "stage": key,
            "at": _now().isoformat(timespec="seconds"),
            "by_id": str(user_id),
            "by_name": await _user_name(session, user_id),
            "note": single_line(clean_text(note))[:2000],
        },
    ]
    await _log(
        session,
        req,
        user_id=user_id,
        what=f"Stage → {spec['name']}",
        detail=single_line(note)[:2000],
    )
    from app.modules.work_requests import events, notifying

    events.stage_changed(req, previous)
    if spec.get("closes"):
        if req.status != "complete":
            # Through the one status path so history and the bell agree -
            # forced, because a closing stage completes from ANY open state.
            await _set_status(
                session, req, "complete", user_id=user_id, can_manage=can_manage, note=note, force=True, dept=dept
            )
        else:
            await notifying.stage_changed(session, req, previous, actor_id=str(user_id))
    else:
        if req.status in ("submitted", "accepted", "on_hold", "complete"):
            # The first move on the board IS the department accepting and
            # starting; a move off a closing stage reopens.
            await _set_status(
                session,
                req,
                "in_progress",
                user_id=user_id,
                can_manage=can_manage,
                note=note,
                force=True,
                dept=dept,
            )
        await notifying.stage_changed(session, req, previous, actor_id=str(user_id))
    await session.flush()
    return req


async def assign(
    session: AsyncSession,
    req: WorkRequest,
    *,
    assignee_ids: list[str] | None,
    responsible_user_id: str | None,
    user_id: str,
    can_manage: bool,
    _checked: bool = False,
) -> WorkRequest:
    dept = await department_or_error(session, req.department)
    if not _checked:
        _require_department(req, dept, user_id, can_manage)
    if req.status in TERMINAL_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status}")
    new_assignees = await _validate_users(session, assignee_ids, field="assignee_ids")
    responsible = (
        (await _validate_users(session, [responsible_user_id], field="responsible_user_id"))[0]
        if responsible_user_id
        else None
    )
    added = [a for a in new_assignees if a not in (req.assignee_ids or [])]
    resp_changed = (responsible or None) != (req.responsible_user_id or None)
    if not added and not resp_changed and set(new_assignees) == set(req.assignee_ids or []):
        return req
    req.assignee_ids = new_assignees
    req.responsible_user_id = responsible
    names = await user_names(session, set(new_assignees) | ({responsible} if responsible else set()))
    detail = ", ".join(names.get(a, a) for a in new_assignees) or "nobody"
    if responsible:
        detail += f"; responsible: {names.get(responsible, responsible)}"
    await _log(session, req, user_id=user_id, what="Assigned", detail=detail)
    from app.modules.work_requests import notifying

    to_tell = list(added)
    if resp_changed and responsible and responsible not in to_tell:
        to_tell.append(responsible)
    await notifying.assigned(session, req, to_tell, actor_id=str(user_id))
    await session.flush()
    return req


async def needs_info(
    session: AsyncSession, req: WorkRequest, question: str, *, user_id: str, can_manage: bool
) -> WorkRequest:
    """The department hands the ball back with a question."""
    dept = await department_or_error(session, req.department)
    _require_department(req, dept, user_id, can_manage)
    if req.status in DONE_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status}")
    text = clean_text(str(question or "")).strip()
    if not text:
        raise WorkRequestError("Say what you need to know")
    req.ball_in_court = "requester"
    req.needs_info = text[:MAX_COMMENT_CHARS]
    await _log(session, req, user_id=user_id, what=text[:MAX_COMMENT_CHARS], kind="needs_info")
    from app.modules.work_requests import events, notifying

    await notifying.needs_info(session, req, text, actor_id=str(user_id))
    events.needs_info(req, text)
    await session.flush()
    return req


async def answer(
    session: AsyncSession, req: WorkRequest, answer_text: str, *, user_id: str, can_manage: bool
) -> WorkRequest:
    """The requester answers; the ball goes back to the department."""
    if req.status in DONE_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status}")
    if req.ball_in_court != "requester" or not req.needs_info:
        raise ConflictError(f"{req.reference} has no open question to answer")
    text = clean_text(str(answer_text or "")).strip()
    if not text:
        raise WorkRequestError("An answer needs some words")
    req.ball_in_court = "department"
    req.needs_info = None
    await _log(session, req, user_id=user_id, what=text[:MAX_COMMENT_CHARS], kind="answer")
    from app.modules.work_requests import events, notifying

    dept = await department_or_error(session, req.department)
    await notifying.answered(session, req, dept, text, actor_id=str(user_id))
    events.answered(req, text)
    await session.flush()
    return req


async def handoff(
    session: AsyncSession,
    req: WorkRequest,
    *,
    department: str,
    request_type: str = "",
    request_types: list[str] | None = None,
    title: str | None = None,
    description: str | None = None,
    due_date: Any = None,
    info_required_by: Any = None,
    copy_links: bool = True,
    user_id: str,
    can_manage: bool,
) -> WorkRequest:
    """Hand part of this work to another department as a CHILD request.

    The child carries ``parent_id``; the parent gains the child in its
    ``depends_on`` - a switchboard build waits on the drafting it asked
    for, and the board shows exactly that.
    """
    dept = await department_or_error(session, req.department)
    _require_update(req, dept, user_id, can_manage)
    if req.status in TERMINAL_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status}")
    target = await department_or_error(session, department)
    child = await create_request(
        session,
        project_id=req.project_id,
        department=target.key,
        request_type=request_type,
        request_types=request_types,
        title=(title or "").strip() or req.title,
        description=description if description is not None else req.description,
        due_date=due_date,
        info_required_by=info_required_by,
        priority=req.priority,
        links=list(req.links or []) if copy_links else [],
        parent_id=req.id,
        user_id=user_id,
        notify=False,
    )
    req.depends_on_ids = [*list(req.depends_on_ids or []), str(child.id)]
    await _log(
        session, req, user_id=user_id, what=f"Handed off to {target.name}", detail=f"{child.reference} - {child.title}"
    )
    await _log(
        session, child, user_id=user_id, what=f"Handed off from {dept.name}", detail=f"{req.reference} - {req.title}"
    )
    from app.modules.work_requests import events, notifying

    await notifying.handoff(session, req, child, target, actor_id=str(user_id))
    events.handoff(req, child)
    await session.flush()
    return child


# ── Duplicating and templating ───────────────────────────────────────────


async def duplicate_request(
    session: AsyncSession,
    src: WorkRequest,
    *,
    title: str | None = None,
    project_id: Any = None,
    user_id: str,
) -> WorkRequest:
    """Copy a request - or a template - into a NEW ``draft``.

    What travels: the request types, the typed fields, the quoted hours,
    who is on it, the links, and the checklist with every item UNTICKED.
    What does not: the hours logged, the conversation, the attachments,
    the stage history and the dependencies. Those are the history of the
    request that was worked, and a fresh copy has none of it yet.

    The copy is a draft on purpose: a duplicate is a starting point
    somebody reads before it goes to the department.
    """
    dept = await department_or_error(session, src.department)
    target_project = src.project_id
    if project_id:
        parsed = _as_uuid(project_id)
        if parsed is None:
            raise WorkRequestError(f"project_id must be a job id, not {project_id!r}")
        target_project = parsed
    new_title = single_line(clean_text(title or ""))[:MAX_TITLE_CHARS] or src.title
    copy = await create_request(
        session,
        project_id=target_project,
        department=dept.key,
        request_types=type_keys_of(src),
        title=new_title,
        description=src.description or "",
        cost_centres=dict(src.cost_centres or {}),
        estimated_hours=dict(src.estimated_hours or {}),
        quoted_hours=src.quoted_hours,
        priority=src.priority,
        links=[dict(x) for x in (src.links or []) if isinstance(x, dict)],
        fields=dict(src.fields or {}),
        assignee_ids=[str(a) for a in (src.assignee_ids or [])],
        responsible_user_id=str(src.responsible_user_id) if src.responsible_user_id else None,
        draft=True,
        user_id=user_id,
        notify=False,
    )
    copy.checklist_state = {}
    # The one-off items travel: they are part of what this request ASKS
    # for, the same as its typed fields. The ticks are what does not.
    copy.checklist_overrides = dict(src.checklist_overrides or {})
    copy.is_template = False
    await _log(session, copy, user_id=user_id, what="Copied from", detail=f"{src.reference} - {src.title}")
    await session.flush()
    return copy


# ── Bulk ─────────────────────────────────────────────────────────────────

#: The only keys a bulk edit may set. Everything else needs the request
#: open in front of you - a title or a set of typed fields is never the
#: same on 40 rows.
BULK_PATCH_KEYS = ("assignee_ids", "responsible_user_id", "stage", "status", "due_date", "priority")


async def bulk_update(
    session: AsyncSession,
    *,
    ids: list[str],
    patch: dict[str, Any],
    user_id: str,
    can_manage: bool,
    project_ids: set[uuid.UUID] | None,
) -> dict[str, Any]:
    """Apply the SAME patch to many requests, one at a time.

    Never all-or-nothing and never silent: each request is applied in its
    own savepoint against the very same rails a single PATCH goes
    through, and one that refuses comes back in ``refused`` with the
    reason while the rest still land.
    """
    wanted: list[str] = []
    for raw in ids or []:
        text = str(raw or "").strip()
        if text and text not in wanted:
            wanted.append(text)
    if not wanted:
        raise WorkRequestError("No requests given")
    if len(wanted) > MAX_BULK_IDS:
        raise WorkRequestError(f"At most {MAX_BULK_IDS} requests in one bulk update, not {len(wanted)}")
    changes = {k: v for k, v in (patch or {}).items()}
    unknown = [k for k in changes if k not in BULK_PATCH_KEYS]
    if unknown:
        raise WorkRequestError(f"{unknown[0]!r} cannot be set in bulk (one of {', '.join(BULK_PATCH_KEYS)})")
    if not changes:
        raise WorkRequestError("Nothing to change")

    updated: list[str] = []
    refused: list[dict[str, str]] = []
    for rid in wanted:
        try:
            async with session.begin_nested():
                req = await request_or_error(session, rid)
                if project_ids is not None and req.project_id not in project_ids:
                    # The same answer for missing and for forbidden, so a
                    # bulk call is never a way to enumerate other jobs.
                    raise NotFoundError("Work request not found")
                await update_request(session, req, dict(changes), user_id=user_id, can_manage=can_manage)
                done_id = str(req.id)
            updated.append(done_id)
        except WorkRequestError as exc:
            refused.append({"id": rid, "reason": str(exc)})
    await session.flush()
    return {"updated": updated, "refused": refused}


# ── Hours ────────────────────────────────────────────────────────────────


async def log_hours(
    session: AsyncSession,
    req: WorkRequest,
    *,
    day: Any,
    hours: Any,
    note: str = "",
    for_user_id: str | None = None,
    user_id: str,
    can_manage: bool,
) -> WorkRequestHours:
    dept = await department_or_error(session, req.department)
    _require_update(req, dept, user_id, can_manage)
    if req.status in TERMINAL_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status}")
    target = str(for_user_id or user_id)
    if target != str(user_id) and not may_act_for_department(req, dept, user_id, can_manage):
        raise NotPermitted("Only the department (or a manager) logs hours for somebody else")
    target = (await _validate_users(session, [target], field="user_id"))[0]
    amount = _hours(hours, field="hours", allow_none=False)
    if not amount:
        raise WorkRequestError("Hours must be more than zero")
    if amount > 24:
        raise WorkRequestError("More than 24 hours in one day is not a day")
    row = WorkRequestHours(
        request_id=req.id,
        user_id=target,
        user_name=await _user_name(session, target),
        day=_iso(day, field="date") or _today().isoformat(),
        hours=amount,
        note=clean_text(str(note or ""))[:2000],
    )
    session.add(row)
    await _log(session, req, user_id=user_id, what=f"Logged {amount:g}h", detail=f"{row.user_name} on {row.day}")
    await session.flush()
    return row


async def list_hours(session: AsyncSession, req: WorkRequest) -> list[dict[str, Any]]:
    rows = (
        await session.execute(
            select(WorkRequestHours)
            .where(WorkRequestHours.request_id == req.id)
            .order_by(WorkRequestHours.day.desc(), WorkRequestHours.created_at.desc())
        )
    ).scalars()
    return [
        {
            "id": str(r.id),
            "user_id": r.user_id,
            "user_name": r.user_name,
            "date": r.day,
            "hours": r.hours,
            "note": r.note or "",
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


async def delete_hours(
    session: AsyncSession, req: WorkRequest, log_id: uuid.UUID, *, user_id: str, can_manage: bool
) -> None:
    row = await session.get(WorkRequestHours, log_id)
    if row is None or row.request_id != req.id:
        raise NotFoundError("Hours entry not found")
    dept = await department_or_error(session, req.department)
    if row.user_id != str(user_id) and not may_act_for_department(req, dept, user_id, can_manage):
        raise NotPermitted("Only the person who logged it, the department or a manager removes an hours entry")
    await _log(session, req, user_id=user_id, what=f"Removed {row.hours:g}h", detail=f"{row.user_name} on {row.day}")
    await session.delete(row)
    await session.flush()


# ── Comments ─────────────────────────────────────────────────────────────


async def add_comment(
    session: AsyncSession, req: WorkRequest, *, body: str, mention_ids: list[str] | None, user_id: str
) -> WorkRequestComment:
    text = clean_text(str(body or "")).strip()
    if not text:
        raise WorkRequestError("A comment needs some words")
    mentions = await _validate_users(session, mention_ids, field="mention_ids")
    row = await _log(session, req, user_id=user_id, what=text[:MAX_COMMENT_CHARS], kind="comment", mention_ids=mentions)
    if mentions:
        from app.modules.work_requests import events, notifying

        await notifying.mentioned(session, req, mentions, actor_id=str(user_id))
        events.mention(req, mentions)
    return row


def comment_payload(c: WorkRequestComment) -> dict[str, Any]:
    return {
        "id": str(c.id),
        "author_id": c.author_id,
        "author_name": c.author_name,
        "body": c.body,
        "detail": c.detail or "",
        "mention_ids": list(c.mention_ids or []),
        "kind": c.kind,
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


async def list_comments(session: AsyncSession, req: WorkRequest, *, include_system: bool = True) -> list[dict]:
    stmt = select(WorkRequestComment).where(WorkRequestComment.request_id == req.id)
    if not include_system:
        stmt = stmt.where(WorkRequestComment.kind != "system")
    stmt = stmt.order_by(WorkRequestComment.created_at)
    return [comment_payload(c) for c in (await session.execute(stmt)).scalars()]


_ACTIVITY_WHAT = {"comment": "Commented", "needs_info": "Asked the requester", "answer": "Answered"}


async def activity(session: AsyncSession, req: WorkRequest) -> list[dict[str, Any]]:
    """``[{at, by_name, what, detail}]``, oldest first."""
    out = []
    for c in await list_comments(session, req):
        if c["kind"] == "system":
            out.append({"at": c["created_at"], "by_name": c["author_name"], "what": c["body"], "detail": c["detail"]})
        else:
            out.append(
                {
                    "at": c["created_at"],
                    "by_name": c["author_name"],
                    "what": _ACTIVITY_WHAT.get(c["kind"], c["kind"]),
                    "detail": c["body"],
                }
            )
    return out


# ── Attachments ──────────────────────────────────────────────────────────


def record_attachment(
    req: WorkRequest, *, filename: str, size: int, mime: str, user_id: str, user_name: str
) -> dict[str, Any]:
    entry = {
        "filename": filename,
        "size": int(size),
        "mime": mime or "application/octet-stream",
        "uploaded_at": _now().isoformat(timespec="seconds"),
        "by_id": str(user_id),
        "by_name": user_name,
    }
    req.attachments = [*list(req.attachments or []), entry]
    return entry


# ── Reading ──────────────────────────────────────────────────────────────


async def request_or_error(session: AsyncSession, id_or_reference: str) -> WorkRequest:
    """By id, or by reference (``WR-ENG-000001``, any case)."""
    rid = _as_uuid(id_or_reference)
    if rid is not None:
        row = await session.get(WorkRequest, rid)
    else:
        row = (
            await session.execute(
                select(WorkRequest).where(func.upper(WorkRequest.reference) == str(id_or_reference).strip().upper())
            )
        ).scalar_one_or_none()
    if row is None:
        raise NotFoundError("Work request not found")
    return row


def _requests_filtered(
    *,
    project_ids: set[uuid.UUID] | None,
    project_id: uuid.UUID | None = None,
    department: str | None = None,
    request_type: str | None = None,
    request_types: str | None = None,
    status: str | None = None,
    stage: str | None = None,
    assignee_id: str | None = None,
    raised_by: str | None = None,
    q: str | None = None,
    include_closed: bool = False,
    is_template: bool | None = None,
) -> Select[tuple[WorkRequest]] | None:
    """The filtered query, before ordering or paging, so the page and the
    count that describes it can never drift apart. ``None`` means the
    caller can see no project at all, which is an empty answer rather than
    an unfiltered one.

    ``project_ids`` is the caller's visibility (None = unrestricted);
    ``project_id`` narrows within it.

    ``is_template`` defaults to False, NOT to "either": a template is a
    thing kept to copy from and never belongs in a register, a board, the
    planner, the summary or the sweep. Ask for ``True`` to list them.
    """
    stmt = select(WorkRequest).where(
        WorkRequest.is_template.is_(True) if is_template else WorkRequest.is_template.is_not(True)
    )
    if project_id is not None:
        stmt = stmt.where(WorkRequest.project_id == project_id)
    elif project_ids is not None:
        if not project_ids:
            return None
        stmt = stmt.where(WorkRequest.project_id.in_(project_ids))
    if department:
        stmt = stmt.where(WorkRequest.department == department.strip().lower())
    wanted_types = [t.strip().lower() for t in f"{request_type or ''},{request_types or ''}".split(",") if t.strip()]
    if wanted_types:
        # MEMBERSHIP, not equality: a request that asks for FDS *and* PLC
        # programming belongs in both of those columns. The singular is
        # still matched so a row written before the list existed is found.
        stmt = stmt.where(
            or_(
                *[
                    or_(
                        WorkRequest.request_type == key,
                        cast(WorkRequest.request_types, String).like("%" + _like_escape(f'"{key}"') + "%", escape="\\"),
                    )
                    for key in wanted_types
                ]
            )
        )
    if status:
        wanted = [s.strip().lower() for s in status.split(",") if s.strip()]
        stmt = stmt.where(WorkRequest.status.in_(wanted))
    elif not include_closed:
        stmt = stmt.where(WorkRequest.status.notin_(list(TERMINAL_STATUSES)))
    if stage:
        stmt = stmt.where(WorkRequest.stage == stage.strip().lower())
    if assignee_id and _as_uuid(assignee_id) is not None:
        stmt = stmt.where(cast(WorkRequest.assignee_ids, String).like(f'%"{str(_as_uuid(assignee_id))}"%'))
    if raised_by:
        stmt = stmt.where(WorkRequest.raised_by_id == str(raised_by).strip())
    if q and q.strip():
        pattern = "%" + _like_escape(q.strip()) + "%"
        stmt = stmt.where(
            or_(
                WorkRequest.reference.ilike(pattern, escape="\\"),
                WorkRequest.title.ilike(pattern, escape="\\"),
                WorkRequest.description.ilike(pattern, escape="\\"),
            )
        )
    return stmt


async def list_requests(
    session: AsyncSession,
    *,
    project_ids: set[uuid.UUID] | None,
    limit: int = 200,
    offset: int = 0,
    **filters: Any,
) -> list[WorkRequest]:
    """One page of the register, newest first."""
    stmt = _requests_filtered(project_ids=project_ids, **filters)
    if stmt is None:
        return []
    stmt = stmt.order_by(WorkRequest.created_at.desc()).limit(max(1, min(limit, 1000))).offset(max(0, offset))
    return list((await session.execute(stmt)).scalars().all())


async def count_requests(
    session: AsyncSession,
    *,
    project_ids: set[uuid.UUID] | None,
    **filters: Any,
) -> int:
    """How many rows the SAME filters match, ignoring the page window."""
    stmt = _requests_filtered(project_ids=project_ids, **filters)
    if stmt is None:
        return 0
    return int((await session.execute(select(func.count()).select_from(stmt.subquery()))).scalar_one())


def _lite(r: WorkRequest) -> dict[str, Any]:
    return {"id": str(r.id), "reference": r.reference, "title": r.title, "status": r.status, "department": r.department}


async def _client_names(session: AsyncSession, client_ids: set[str]) -> dict[str, str]:
    """Resolve ``Project.client_id`` the way the standup board does: a UUID
    names a contact, anything else is the client's name typed verbatim."""
    out: dict[str, str] = {}
    wanted = {}
    for raw in client_ids:
        text = str(raw or "").strip()
        if not text:
            continue
        if _as_uuid(text) is None:
            out[text] = text
        else:
            wanted[text.lower()] = text
    if wanted:
        try:
            from app.modules.contacts.models import Contact

            rows = (
                await session.execute(
                    select(Contact.id, Contact.company_name, Contact.first_name, Contact.last_name).where(
                        Contact.id.in_([uuid.UUID(k) for k in wanted])
                    )
                )
            ).all()
            for cid, company, first, last in rows:
                label = (company or "").strip() or " ".join(x for x in (first, last) if x).strip()
                if label:
                    out[wanted[str(cid).lower()]] = label
        except Exception:  # noqa: BLE001 - contacts is optional
            logger.debug("work_requests: client lookup unavailable", exc_info=True)
    return out


async def payloads(session: AsyncSession, requests: list[WorkRequest]) -> list[dict[str, Any]]:
    """Every request as the UI wants it, in a handful of batched queries."""
    if not requests:
        return []
    from app.modules.projects.models import Project

    ids = [r.id for r in requests]
    id_strs = {str(i) for i in ids}
    departments = await _departments_by_key(session)

    project_ids = {r.project_id for r in requests}
    project_rows = (
        await session.execute(
            select(Project.id, Project.name, Project.project_code, Project.client_id).where(Project.id.in_(project_ids))
        )
    ).all()
    projects = {str(pid): (name, code, client) for pid, name, code, client in project_rows}
    clients = await _client_names(session, {c for _n, _c, c in projects.values() if c})

    people: set[str] = set()
    for r in requests:
        people.add(str(r.raised_by_id or ""))
        people.update(str(a) for a in (r.assignee_ids or []))
        if r.responsible_user_id:
            people.add(str(r.responsible_user_id))
    names = await user_names(session, {p for p in people if p})

    hours_rows = (
        await session.execute(
            select(WorkRequestHours.request_id, func.sum(WorkRequestHours.hours))
            .where(WorkRequestHours.request_id.in_(ids))
            .group_by(WorkRequestHours.request_id)
        )
    ).all()
    hours_logged = {str(rid): float(total or 0) for rid, total in hours_rows}
    comment_rows = (
        await session.execute(
            select(WorkRequestComment.request_id, func.count())
            .where(WorkRequestComment.request_id.in_(ids), WorkRequestComment.kind != "system")
            .group_by(WorkRequestComment.request_id)
        )
    ).all()
    comment_count = {str(rid): int(n) for rid, n in comment_rows}

    # Relatives: what each depends on, what depends on each, parents, children.
    related_ids: set[uuid.UUID] = set()
    for r in requests:
        for d in r.depends_on_ids or []:
            u = _as_uuid(d)
            if u is not None:
                related_ids.add(u)
        if r.parent_id:
            related_ids.add(r.parent_id)
    related: dict[str, WorkRequest] = {str(r.id): r for r in requests}
    missing = [i for i in related_ids if str(i) not in related]
    if missing:
        for row in (await session.execute(select(WorkRequest).where(WorkRequest.id.in_(missing)))).scalars():
            related[str(row.id)] = row
    blockers_stmt = select(WorkRequest).where(
        or_(*[cast(WorkRequest.depends_on_ids, String).like(f'%"{i}"%') for i in id_strs])
    )
    blocks: dict[str, list[WorkRequest]] = {i: [] for i in id_strs}
    for row in (await session.execute(blockers_stmt)).scalars():
        for d in row.depends_on_ids or []:
            if str(d) in blocks:
                blocks[str(d)].append(row)
    children: dict[str, list[WorkRequest]] = {i: [] for i in id_strs}
    for row in (await session.execute(select(WorkRequest).where(WorkRequest.parent_id.in_(ids)))).scalars():
        children[str(row.parent_id)].append(row)

    today = _today()
    out: list[dict[str, Any]] = []
    for r in requests:
        dept = departments.get(r.department)
        pname, pcode, pclient = projects.get(str(r.project_id), ("", None, None))
        logged = round(hours_logged.get(str(r.id), 0.0), 2)
        to_complete = r.hours_to_complete
        at_completion = round(logged + (to_complete or 0.0), 2)
        deviation = round(at_completion - r.quoted_hours, 2) if r.quoted_hours is not None else None
        cost = None
        if dept is not None and dept.hourly_rate:
            try:
                cost = f"{(Decimal(str(at_completion)) * Decimal(dept.hourly_rate)).quantize(Decimal('0.01'))}"
            except (InvalidOperation, ValueError):
                cost = None
        days_until_due = None
        is_overdue = False
        if r.due_date:
            try:
                days_until_due = (date.fromisoformat(r.due_date) - today).days
                is_overdue = days_until_due < 0 and r.status in ACTIVE_STATUSES
            except ValueError:
                pass
        stage = stage_spec(dept, r.stage) if dept else None
        parent = related.get(str(r.parent_id)) if r.parent_id else None
        # Backfilled on READ: a row written before the column existed
        # carries only the singular, and the UI must never have to care.
        rtype_keys = type_keys_of(r)
        rtypes = specs_on(dept, rtype_keys)
        labels = {t["key"]: (t.get("label") or t["key"]) for t in rtypes}
        checklist = checklist_for(r, dept)
        target_date, days_late, is_late = lateness(r, dept, today=today)
        out.append(
            {
                "id": str(r.id),
                "reference": r.reference,
                "project_id": str(r.project_id),
                "project_code": pcode or "",
                "project_name": pname or "",
                "client_name": clients.get(str(pclient or ""), "") if pclient else "",
                "department": r.department,
                "department_name": dept.name if dept else r.department,
                "request_type": r.request_type,
                "request_types": rtype_keys,
                "request_type_labels": [labels.get(k, k) for k in rtype_keys],
                "field_specs": merged_field_specs(rtypes),
                "title": r.title,
                "description": r.description or "",
                "status": r.status,
                "stage": r.stage,
                "stage_name": stage["name"] if stage else None,
                "stage_closes": bool(stage.get("closes")) if stage else False,
                "stage_history": list(r.stage_history or []),
                "raised_by_id": r.raised_by_id,
                "raised_by_name": r.raised_by_name or names.get(r.raised_by_id, ""),
                "assignees": [{"id": a, "name": names.get(a, "")} for a in (r.assignee_ids or [])],
                "responsible": (
                    {"id": r.responsible_user_id, "name": names.get(r.responsible_user_id, "")}
                    if r.responsible_user_id
                    else None
                ),
                "cost_centres": dict(r.cost_centres or {}),
                "estimated_hours": dict(r.estimated_hours or {}),
                "quoted_hours": r.quoted_hours,
                "hours_logged": logged,
                "hours_to_complete": to_complete,
                "hours_at_completion": at_completion,
                "deviation_hours": deviation,
                "cost_at_completion": cost,
                "checklist": checklist,
                "checklist_done": sum(1 for i in checklist if i["done"]),
                "checklist_total": len(checklist),
                # The DIFFERENCE from the type's list, not a copy of it -
                # what a "reset to the request type" button greys out.
                "checklist_overrides": overrides_of(r),
                "checklist_is_overridden": bool(r.checklist_overrides or {}),
                "info_required_by": r.info_required_by,
                "due_date": r.due_date,
                "days_until_due": days_until_due,
                "is_overdue": is_overdue,
                "accepted_at": r.accepted_at,
                "target_days": dept.target_days if dept is not None else None,
                "target_date": target_date,
                "days_late": days_late,
                "is_late": is_late,
                "is_template": bool(r.is_template),
                "schedule_activity_id": str(r.schedule_activity_id) if r.schedule_activity_id else None,
                "boq_position_ids": [str(p) for p in (r.boq_position_ids or [])],
                "scheduled_start": r.scheduled_start,
                "scheduled_end": r.scheduled_end,
                "delivered_at": r.delivered_at,
                "tested_at": r.tested_at,
                "priority": r.priority,
                "links": list(r.links or []),
                "fields": dict(r.fields or {}),
                "planner_uploaded": bool(r.planner_uploaded),
                "ball_in_court": r.ball_in_court if r.ball_in_court in BALL_IN_COURT else "department",
                "needs_info": r.needs_info,
                "depends_on": [_lite(related[str(d)]) for d in (r.depends_on_ids or []) if str(d) in related],
                "blocks": [_lite(b) for b in blocks.get(str(r.id), [])],
                "parent_id": str(r.parent_id) if r.parent_id else None,
                "parent_reference": parent.reference if parent else None,
                "children": [_lite(c) for c in children.get(str(r.id), [])],
                "comment_count": comment_count.get(str(r.id), 0),
                "attachments": [
                    {"filename": a.get("filename"), "size": a.get("size", 0), "uploaded_at": a.get("uploaded_at")}
                    for a in (r.attachments or [])
                    if isinstance(a, dict)
                ],
                "allowed_transitions": allowed_transitions(r),
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
                "closed_at": r.closed_at.isoformat() if r.closed_at else None,
            }
        )
    return out


async def payload(session: AsyncSession, req: WorkRequest) -> dict[str, Any]:
    return (await payloads(session, [req]))[0]


# ── Export ───────────────────────────────────────────────────────────────

#: The columns the exported sheet has, in its established order. The order
#: IS the contract - somebody's macro reads column R - so new columns are
#: appended, never inserted.
EXPORT_COLUMNS: tuple[str, ...] = (
    "Reference",
    "Department",
    "Request types",
    "Title",
    "Job number",
    "Client",
    "Raised by",
    "Raised on",
    "Assignees",
    "Responsible",
    "Stage",
    "Status",
    "Ball in court",
    "Due date",
    "Info required by",
    "Quoted hours",
    "Logged hours",
    "Hours to complete",
    "Hours at completion",
    "Deviation",
    "Hourly rate",
    "Cost at completion",
    "Checklist",
    "Last activity",
)

#: A cell a spreadsheet would run instead of showing. Prefixed with an
#: apostrophe so Excel renders the text - the export is a report, and a
#: report must never be an execution path.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def _formula_safe(text: str) -> str:
    if text and text[0] in _FORMULA_LEAD:
        try:
            Decimal(text)  # a plain number is a number, not a formula
        except InvalidOperation:
            return "'" + text
    return text


def _number_text(value: Any) -> str:
    """Hours and quantities as plain decimal text - never a float in a
    cell, the same convention money already has here."""
    if value is None or value == "":
        return ""
    try:
        return format(Decimal(str(round(float(value), 2))).quantize(Decimal("0.01")), "f")
    except (InvalidOperation, TypeError, ValueError):
        return ""


def export_row(p: dict[str, Any], *, hourly_rate: str | None) -> list[str]:
    """One request as its export line, every cell already text."""
    raised_on = str(p.get("created_at") or "")[:10]
    cells = [
        p.get("reference") or "",
        p.get("department_name") or p.get("department") or "",
        "; ".join(p.get("request_type_labels") or []),
        p.get("title") or "",
        p.get("project_code") or "",
        p.get("client_name") or "",
        p.get("raised_by_name") or "",
        raised_on,
        "; ".join(a.get("name") or a.get("id") or "" for a in (p.get("assignees") or [])),
        (p.get("responsible") or {}).get("name") or "" if p.get("responsible") else "",
        p.get("stage_name") or p.get("stage") or "",
        p.get("status") or "",
        p.get("ball_in_court") or "",
        p.get("due_date") or "",
        p.get("info_required_by") or "",
        _number_text(p.get("quoted_hours")),
        _number_text(p.get("hours_logged")),
        _number_text(p.get("hours_to_complete")),
        _number_text(p.get("hours_at_completion")),
        _number_text(p.get("deviation_hours")),
        hourly_rate or "",
        p.get("cost_at_completion") or "",
        f"{p.get('checklist_done', 0)}/{p.get('checklist_total', 0)}",
        p.get("updated_at") or "",
    ]
    return [_formula_safe(str(c)) for c in cells]


async def export_table(session: AsyncSession, requests: list[WorkRequest]) -> list[list[str]]:
    """The header row followed by one row per request."""
    departments = await _departments_by_key(session)
    rows = [list(EXPORT_COLUMNS)]
    for p in await payloads(session, requests):
        dept = departments.get(p.get("department") or "")
        rows.append(export_row(p, hourly_rate=dept.hourly_rate if dept else None))
    return rows


def export_csv_bytes(rows: list[list[str]]) -> bytes:
    """RFC-4180 CSV, CRLF, UTF-8 WITH a BOM - the export opens in Excel
    by double-click and a name with an accent survives it."""
    import csv
    import io

    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx_bytes(rows: list[list[str]]) -> bytes | None:
    """A real workbook, or None when no spreadsheet library is installed -
    the caller falls back to CSV rather than this module growing a
    dependency behind an operator's back."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:  # pragma: no cover - openpyxl is a project dependency
        return None
    import io

    book = Workbook()
    sheet = book.active
    sheet.title = "Work requests"
    for row in rows:
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    for i, header in enumerate(rows[0] if rows else [], start=1):
        sheet.column_dimensions[get_column_letter(i)].width = max(12, min(48, len(str(header)) + 6))
    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def export_filename(department: str | None, fmt: str) -> str:
    """``work-requests-workshop-2026-09-03.csv`` - the department it was
    taken for, or ``all``."""
    slug = re.sub(r"[^a-z0-9_-]+", "-", str(department or "").strip().lower()).strip("-") or "all"
    return f"work-requests-{slug}-{_today().isoformat()}.{fmt}"


async def export_requests(
    session: AsyncSession,
    *,
    fmt: str = "csv",
    department: str | None = None,
    **filters: Any,
) -> dict[str, Any]:
    """The register as a file. Accepts exactly the filters ``list_requests``
    does, so what somebody exports is what they were looking at.

    Returns ``{filename, media_type, content, rows, format, note}``.
    ``note`` is set only when a requested ``xlsx`` had to fall back.
    """
    wanted = str(fmt or "csv").strip().lower()
    if wanted not in ("csv", "xlsx"):
        raise WorkRequestError(f"Unknown export format {fmt!r} (csv or xlsx)")
    collected: list[WorkRequest] = []
    page = 1000
    for offset in range(0, MAX_EXPORT_ROWS, page):
        batch = await list_requests(session, department=department, limit=page, offset=offset, **filters)
        collected.extend(batch)
        if len(batch) < page:
            break
    rows = await export_table(session, collected[:MAX_EXPORT_ROWS])
    note = ""
    content: bytes | None = None
    if wanted == "xlsx":
        content = export_xlsx_bytes(rows)
        if content is None:
            wanted = "csv"
            note = "No spreadsheet library is installed on this deployment - exported as CSV instead"
    if content is None:
        content = export_csv_bytes(rows)
    return {
        "filename": export_filename(department, wanted),
        "media_type": (
            "text/csv; charset=utf-8"
            if wanted == "csv"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "content": content,
        "rows": len(rows) - 1,
        "format": wanted,
        "note": note,
    }


# ── Planner ──────────────────────────────────────────────────────────────

PLANNER_DEFAULT_DAYS = 35
PLANNER_MAX_DAYS = 120


def _day_range(from_day: Any, to_day: Any) -> list[str]:
    start = date.fromisoformat(_iso(from_day, field="from") or _today().isoformat())
    end_text = _iso(to_day, field="to")
    end = date.fromisoformat(end_text) if end_text else start + timedelta(days=PLANNER_DEFAULT_DAYS - 1)
    if end < start:
        raise WorkRequestError("'to' is before 'from'")
    if (end - start).days >= PLANNER_MAX_DAYS:
        raise WorkRequestError(f"The planner shows at most {PLANNER_MAX_DAYS} days at a time")
    return [(start + timedelta(days=i)).isoformat() for i in range((end - start).days + 1)]


async def planner(session: AsyncSession, *, department: str, from_day: Any = None, to_day: Any = None) -> dict:
    """The headcount grid: every open request in the department against
    every day in the window, with the capacity line underneath."""
    dept = await department_or_error(session, department)
    days = _day_range(from_day, to_day)
    member_ids = [str(m) for m in (dept.member_ids or [])]
    if dept.lead_user_id and str(dept.lead_user_id) not in member_ids:
        member_ids.insert(0, str(dept.lead_user_id))
    names = await user_names(session, set(member_ids))
    members = [{"id": m, "name": names.get(m, "")} for m in member_ids if m in names]

    rows = (
        (
            await session.execute(
                select(WorkRequest)
                .where(
                    WorkRequest.department == dept.key,
                    WorkRequest.status.in_(list(ACTIVE_STATUSES)),
                    WorkRequest.is_template.is_not(True),
                )
                .order_by(WorkRequest.due_date.nulls_last(), WorkRequest.created_at)
            )
        )
        .scalars()
        .all()
    )
    ids = [r.id for r in rows]
    allocs: dict[str, dict[str, float]] = {str(i): {} for i in ids}
    if ids:
        for a in (
            await session.execute(
                select(WorkPlannerAlloc).where(
                    WorkPlannerAlloc.request_id.in_(ids),
                    WorkPlannerAlloc.day >= days[0],
                    WorkPlannerAlloc.day <= days[-1],
                )
            )
        ).scalars():
            allocs[str(a.request_id)][a.day] = float(a.people)
    caps = {
        c.day: float(c.available)
        for c in (
            await session.execute(
                select(WorkPlannerCapacity).where(
                    WorkPlannerCapacity.department == dept.key,
                    WorkPlannerCapacity.day >= days[0],
                    WorkPlannerCapacity.day <= days[-1],
                )
            )
        ).scalars()
    }
    all_names = await user_names(session, {a for r in rows for a in (r.assignee_ids or [])})
    stages = {s["key"]: s for s in (dept.stages or [])}
    out_rows = []
    for r in rows:
        out_rows.append(
            {
                "request_id": str(r.id),
                "reference": r.reference,
                "title": r.title,
                "project_code": "",
                "due_date": r.due_date,
                "status": r.status,
                "stage": r.stage,
                "stage_name": (stages.get(r.stage) or {}).get("name"),
                "assignees": [{"id": a, "name": all_names.get(a, "")} for a in (r.assignee_ids or [])],
                "alloc": allocs.get(str(r.id), {}),
            }
        )
    # Project codes in one query rather than one per row.
    from app.modules.projects.models import Project

    codes = (
        {
            str(pid): code or ""
            for pid, code in (
                await session.execute(
                    select(Project.id, Project.project_code).where(Project.id.in_({r.project_id for r in rows}))
                )
            ).all()
        }
        if rows
        else {}
    )
    for r, row in zip(rows, out_rows, strict=True):
        row["project_code"] = codes.get(str(r.project_id), "")

    capacity = {}
    for day in days:
        allocated = round(sum(allocs[str(r.id)].get(day, 0.0) for r in rows), 2)
        capacity[day] = {
            "available": caps.get(day, float(len(members))),
            "allocated": allocated,
            "override": day in caps,
        }
    return {
        "department": dept.key,
        "from": days[0],
        "to": days[-1],
        "days": days,
        "members": members,
        "rows": out_rows,
        "capacity": capacity,
    }


async def set_allocation(
    session: AsyncSession, req: WorkRequest, alloc: dict[str, Any], *, user_id: str, can_manage: bool
) -> dict[str, float]:
    """Upsert people-per-day for one request. Zero or null clears a day."""
    dept = await department_or_error(session, req.department)
    _require_department(req, dept, user_id, can_manage)
    if req.status in TERMINAL_STATUSES:
        raise ConflictError(f"{req.reference} is {req.status}")
    wanted: dict[str, float | None] = {}
    for raw_day, raw_people in (alloc or {}).items():
        day = _iso(raw_day, field="day")
        if day is None:
            continue
        people = _hours(raw_people, field=f"people on {day}")
        wanted[day] = people
    if not wanted:
        raise WorkRequestError("Nothing to allocate")
    existing = {
        a.day: a
        for a in (
            await session.execute(
                select(WorkPlannerAlloc).where(
                    WorkPlannerAlloc.request_id == req.id, WorkPlannerAlloc.day.in_(list(wanted))
                )
            )
        ).scalars()
    }
    for day, people in wanted.items():
        row = existing.get(day)
        if not people:
            if row is not None:
                await session.delete(row)
            continue
        if row is None:
            session.add(WorkPlannerAlloc(request_id=req.id, day=day, people=people))
        else:
            row.people = people
    await session.flush()
    rows = (await session.execute(select(WorkPlannerAlloc).where(WorkPlannerAlloc.request_id == req.id))).scalars()
    return {a.day: float(a.people) for a in rows}


async def set_capacity(session: AsyncSession, department: str, capacity: dict[str, Any]) -> dict[str, float]:
    dept = await department_or_error(session, department)
    wanted: dict[str, float | None] = {}
    for raw_day, raw_people in (capacity or {}).items():
        day = _iso(raw_day, field="day")
        if day is not None:
            wanted[day] = _hours(raw_people, field=f"available on {day}")
    if not wanted:
        raise WorkRequestError("Nothing to set")
    existing = {
        c.day: c
        for c in (
            await session.execute(
                select(WorkPlannerCapacity).where(
                    WorkPlannerCapacity.department == dept.key, WorkPlannerCapacity.day.in_(list(wanted))
                )
            )
        ).scalars()
    }
    for day, people in wanted.items():
        row = existing.get(day)
        if people is None:
            if row is not None:
                await session.delete(row)
            continue
        if row is None:
            session.add(WorkPlannerCapacity(department=dept.key, day=day, available=people))
        else:
            row.available = people
    await session.flush()
    rows = (
        await session.execute(select(WorkPlannerCapacity).where(WorkPlannerCapacity.department == dept.key))
    ).scalars()
    return {c.day: float(c.available) for c in rows}


# ── Summary and my queue ─────────────────────────────────────────────────


async def summary(
    session: AsyncSession, *, project_ids: set[uuid.UUID] | None, project_id: uuid.UUID | None = None
) -> dict:
    """Per department: what is open, what is late, what is waiting on the
    requester, what falls due this week, hours quoted vs logged."""
    await ensure_seeded(session)
    rows = await list_requests(
        session, project_ids=project_ids, project_id=project_id, include_closed=False, limit=1000
    )
    active = [r for r in rows if r.status in ACTIVE_STATUSES]
    logged: dict[str, float] = {}
    if active:
        for rid, total in (
            await session.execute(
                select(WorkRequestHours.request_id, func.sum(WorkRequestHours.hours))
                .where(WorkRequestHours.request_id.in_([r.id for r in active]))
                .group_by(WorkRequestHours.request_id)
            )
        ).all():
            logged[str(rid)] = float(total or 0)
    today = _today()
    week_end = today + timedelta(days=7)
    out = []
    for d in await list_departments(session):
        if not d.active:
            continue
        mine = [r for r in active if r.department == d.key]
        overdue = 0
        due_this_week = 0
        late = sum(1 for r in mine if lateness(r, d, today=today)[2])
        for r in mine:
            if not r.due_date:
                continue
            try:
                due = date.fromisoformat(r.due_date)
            except ValueError:
                continue
            if due < today:
                overdue += 1
            elif due <= week_end:
                due_this_week += 1
        out.append(
            {
                "key": d.key,
                "name": d.name,
                "colour": d.colour,
                "open": len(mine),
                "overdue": overdue,
                #: Past the DEPARTMENT'S OWN turnaround target - a
                #: different question from ``overdue``, which is past the
                #: date the requester asked for.
                "late": late,
                "target_days": d.target_days,
                "with_requester": sum(1 for r in mine if r.ball_in_court == "requester"),
                "due_this_week": due_this_week,
                "hours_quoted": round(sum(r.quoted_hours or 0.0 for r in mine), 2),
                "hours_logged": round(sum(logged.get(str(r.id), 0.0) for r in mine), 2),
                "awaiting_close": sum(1 for r in rows if r.department == d.key and r.status == "complete"),
            }
        )
    return {"departments": out}


async def my_queue(session: AsyncSession, *, user_id: str, project_ids: set[uuid.UUID] | None) -> dict:
    uid = str(user_id)
    rows = await list_requests(session, project_ids=project_ids, include_closed=False, limit=1000)
    active = [r for r in rows if r.status in ACTIVE_STATUSES]
    assigned = [r for r in active if uid in {str(a) for a in (r.assignee_ids or [])}]
    responsible = [r for r in active if str(r.responsible_user_id or "") == uid]
    raised = [r for r in rows if str(r.raised_by_id) == uid]
    needs_my_answer = [
        r for r in raised if r.status in ACTIVE_STATUSES and r.ball_in_court == "requester" and r.needs_info
    ]
    seen: dict[str, WorkRequest] = {}
    for r in [*assigned, *responsible, *raised]:
        seen.setdefault(str(r.id), r)
    built = {p["id"]: p for p in await payloads(session, list(seen.values()))}
    return {
        "assigned": [built[str(r.id)] for r in assigned],
        "responsible": [built[str(r.id)] for r in responsible],
        "raised": [built[str(r.id)] for r in raised],
        "needs_my_answer": [built[str(r.id)] for r in needs_my_answer],
    }
